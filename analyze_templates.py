#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para analizar los templates de 賃金台帳
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def analyze_template(filepath: str, template_name: str):
    print("=" * 100)
    print(f"ANALIZANDO TEMPLATE: {template_name}")
    print("=" * 100)

    wb = load_workbook(filepath, data_only=True)
    print(f"\nHojas disponibles: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        print(f"\n{'=' * 100}")
        print(f"HOJA: {sheet_name}")
        print("=" * 100)

        ws = wb[sheet_name]
        print(f"Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")

        # Mostrar todas las celdas con contenido (primeras 100 filas)
        print(f"\nContenido de celdas (primeras 100 filas):")
        print("-" * 100)

        for row in range(1, min(ws.max_row + 1, 101)):
            row_content = {}
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    col_letter = get_column_letter(col)
                    row_content[col_letter] = str(cell.value)[:50]  # Limitar a 50 chars

            if row_content:
                print(f"Fila {row:3d}: {row_content}")

        # Analizar formato
        print(f"\n{'=' * 100}")
        print("ANÁLISIS DE FORMATO:")
        print("=" * 100)

        # Buscar títulos clave
        keywords = ['賃金台帳', '氏名', '社員番号', '勤務日数', '基本給', '控除', '支給', '差引支給額']
        for keyword in keywords:
            found = False
            for row in range(1, min(ws.max_row + 1, 50)):
                for col in range(1, min(ws.max_column + 1, 30)):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and keyword in str(cell.value):
                        col_letter = get_column_letter(col)
                        print(f"  '{keyword}' encontrado en {col_letter}{row}: {cell.value}")
                        found = True
                        break
                if found:
                    break

    wb.close()
    print("\n")

# Analizar ambos templates
try:
    analyze_template(r"C:\Users\JPUNS\Downloads\wageledgersample.xlsx", "TEMPLATE 1: wageledgersample.xlsx")
except Exception as e:
    print(f"ERROR analizando wageledgersample.xlsx: {e}")
    import traceback
    traceback.print_exc()

print("\n" * 3)

try:
    analyze_template(r"C:\Users\JPUNS\Downloads\tingin01.xlsx", "TEMPLATE 2: tingin01.xlsx")
except Exception as e:
    print(f"ERROR analizando tingin01.xlsx: {e}")
    import traceback
    traceback.print_exc()

print("\n" + "=" * 100)
print("ANÁLISIS COMPLETADO")
print("=" * 100)
