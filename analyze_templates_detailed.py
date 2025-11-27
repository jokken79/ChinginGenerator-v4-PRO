#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analisis detallado de templates para replicar estructura
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

def analyze_template_detailed(filepath: str, template_name: str):
    """Analizar template y extraer toda la estructura"""
    wb = load_workbook(filepath, data_only=False)  # data_only=False para ver formulas

    result = {
        "template_name": template_name,
        "sheets": []
    }

    for sheet_name in wb.sheetnames:
        sheet_data = {
            "name": sheet_name,
            "dimensions": f"{wb[sheet_name].max_row}x{wb[sheet_name].max_column}",
            "cells": [],
            "merged_cells": [],
            "styles": {}
        }

        ws = wb[sheet_name]

        # Obtener celdas combinadas
        for merged in ws.merged_cells.ranges:
            sheet_data["merged_cells"].append(str(merged))

        # Analizar todas las celdas con contenido (primeras 100 filas)
        for row in range(1, min(ws.max_row + 1, 101)):
            for col in range(1, min(ws.max_column + 1, 70)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None or cell.fill.start_color.index != '00000000':
                    col_letter = get_column_letter(col)
                    cell_ref = f"{col_letter}{row}"

                    # Convert color objects to strings
                    font_color = None
                    if cell.font.color:
                        font_color = str(cell.font.color.rgb) if hasattr(cell.font.color, 'rgb') else str(cell.font.color)

                    fill_color = None
                    if cell.fill.start_color:
                        fill_color = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color)

                    cell_info = {
                        "ref": cell_ref,
                        "row": row,
                        "col": col,
                        "value": str(cell.value)[:100] if cell.value else None,
                        "font": {
                            "name": cell.font.name,
                            "size": cell.font.size,
                            "bold": cell.font.bold,
                            "color": font_color
                        },
                        "fill": {
                            "type": cell.fill.fill_type,
                            "color": fill_color
                        },
                        "alignment": {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text
                        },
                        "border": {
                            "top": bool(cell.border.top.style) if cell.border.top else False,
                            "bottom": bool(cell.border.bottom.style) if cell.border.bottom else False,
                            "left": bool(cell.border.left.style) if cell.border.left else False,
                            "right": bool(cell.border.right.style) if cell.border.right else False
                        },
                        "number_format": cell.number_format
                    }

                    sheet_data["cells"].append(cell_info)

        result["sheets"].append(sheet_data)

    wb.close()
    return result

# Analizar ambos templates
print("Analizando templates...")

try:
    print("Template B: wageledgersample.xlsx...")
    result_b = analyze_template_detailed(
        r"d:\ChinginGenerator-v4-PRO\templates\template_format_b.xlsx",
        "Format B - Horizontal 12 Months (Detailed)"
    )

    with open("d:\\ChinginGenerator-v4-PRO\\template_b_analysis.json", "w", encoding="utf-8") as f:
        json.dump(result_b, f, ensure_ascii=False, indent=2)
    print("  OK - Guardado en template_b_analysis.json")

except Exception as e:
    print(f"ERROR Template B: {e}")
    import traceback
    traceback.print_exc()

try:
    print("\nTemplate C: tingin01.xlsx...")
    result_c = analyze_template_detailed(
        r"d:\ChinginGenerator-v4-PRO\templates\template_format_c.xlsx",
        "Format C - Horizontal 12 Months (Simplified)"
    )

    with open("d:\\ChinginGenerator-v4-PRO\\template_c_analysis.json", "w", encoding="utf-8") as f:
        json.dump(result_c, f, ensure_ascii=False, indent=2)
    print("  OK - Guardado en template_c_analysis.json")

except Exception as e:
    print(f"ERROR Template C: {e}")
    import traceback
    traceback.print_exc()

print("\nAnalisis completado!")
print("Revisa los archivos JSON para ver la estructura detallada")
