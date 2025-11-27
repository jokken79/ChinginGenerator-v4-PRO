#!/usr/bin/env python3
"""
Excel Processor v4 PRO para 賃金台帳 Generator
- Lee datos de archivos Excel con formato 給与明細
- Lee directamente de la hoja totalChin que genera la macro
- Exporta con TODAS las columnas originales
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
import os
import re
import json

from database import (
    init_database, save_payroll_record, get_all_payroll_records,
    get_payroll_by_employee, get_payroll_by_period, get_periods,
    get_all_employees, log_audit, check_auto_backup
)


class ExcelProcessor:
    """Procesa archivos de 給与明細 y guarda en base de datos"""
    
    # Headers exactos según la macro VBA (53 columnas)
    HEADERS_FULL = [
        "Number", "従業員番号", "氏名ローマ字", "氏名", "支給分", "派遣先",
        "賃金計算期間S", "賃金計算期間F", "出勤日数", "欠勤日数", "有給日数", "早退時間",
        "実働時", "実働時分", "残業時間数", "残業時間数分", "深夜労働時間数", "深夜労働時間数分",
        "基本給 (時給)", "普通残業手当", "深夜残業手当", "休日勤務手当", "有給休暇",
        "1", "2", "3", "4", "5", "6", "7", "8",
        "前月給与", "合計", "健康保険料", "厚生年金", "雇用保険料", "社会保険料計",
        "住民税", "所得税", "控除1", "控除2", "控除3", "控除4", "控除5", "控除6",
        "控除7", "控除8", "控除9", "控除合計", "差引支給額", "通勤手当(非)",
        "その他手当1", "その他"
    ]
    
    # Índices de columnas importantes (0-based)
    IDX = {
        "number": 0,
        "employee_id": 1,
        "name_roman": 2,
        "name_jp": 3,
        "period": 4,
        "dispatch": 5,
        "period_start": 6,
        "period_end": 7,
        "work_days": 8,
        "absence_days": 9,
        "paid_leave_days": 10,
        "early_leave": 11,
        "work_hours": 12,
        "work_minutes": 13,
        "overtime_hours": 14,
        "overtime_minutes": 15,
        "night_hours": 16,
        "night_minutes": 17,
        "base_pay": 18,
        "overtime_pay": 19,
        "night_pay": 20,
        "holiday_pay": 21,
        "paid_leave_pay": 22,
        "allowance_1": 23,
        "allowance_2": 24,
        "allowance_3": 25,
        "allowance_4": 26,
        "allowance_5": 27,
        "allowance_6": 28,
        "allowance_7": 29,
        "allowance_8": 30,
        "prev_month_pay": 31,
        "total_pay": 32,
        "health_insurance": 33,
        "pension": 34,
        "employment_insurance": 35,
        "social_total": 36,
        "resident_tax": 37,
        "income_tax": 38,
        "deduction_1": 39,
        "deduction_2": 40,
        "deduction_3": 41,
        "deduction_4": 42,
        "deduction_5": 43,
        "deduction_6": 44,
        "deduction_7": 45,
        "deduction_8": 46,
        "deduction_9": 47,
        "deduction_total": 48,
        "net_pay": 49,
        "commuting_allowance": 50,
        "other_allowance_1": 51,
        "other": 52,
    }
    
    def __init__(self):
        self.processed_files = []
        self.errors = []
        self.records_saved = 0
        self.all_records = []  # Almacenar todos los registros con todas las columnas
        
        init_database()
        check_auto_backup()
    
    def process_file(self, filepath: str) -> dict:
        """Procesa un archivo Excel y guarda en BD"""
        filename = os.path.basename(filepath)
        
        try:
            wb = load_workbook(filepath, data_only=True)
            
            # Buscar hoja de datos consolidados
            sheet_name = None
            priority_sheets = ["totalChin", "2025年", "総合", "ALL", "全員"]
            
            for priority in priority_sheets:
                for name in wb.sheetnames:
                    if priority.lower() in name.lower():
                        sheet_name = name
                        break
                if sheet_name:
                    break
            
            if not sheet_name:
                # Buscar cualquier hoja que empiece con año
                for name in wb.sheetnames:
                    if re.match(r'\d{4}年', name):
                        sheet_name = name
                        break
            
            if not sheet_name:
                sheet_name = wb.sheetnames[0]
            
            print(f"   [Procesando] hoja: {sheet_name}")
            
            ws = wb[sheet_name]
            records_count = 0
            
            # Leer headers de la hoja
            headers = []
            for col in range(1, 100):
                h = ws.cell(row=1, column=col).value
                if h is None and col > 53:
                    break
                headers.append(h)
            
            print(f"   [INFO] Columnas encontradas: {len(headers)}")
            
            # Detectar posiciones dinámicas de columnas especiales
            commuting_idx = None
            for idx, h in enumerate(headers):
                if h and '通勤' in str(h) and '非' in str(h):
                    commuting_idx = idx
                    print(f"   [INFO] 通勤手当(非) detectado en columna {idx + 1} (indice {idx})")
                    break
            
            # Procesar cada fila
            for row_idx in range(2, ws.max_row + 1):
                # Leer toda la fila
                row_data = []
                for col in range(1, len(headers) + 1):
                    row_data.append(ws.cell(row=row_idx, column=col).value)
                
                # Validar que tenga employee_id
                employee_id = str(row_data[1]).strip() if len(row_data) > 1 and row_data[1] else None
                if not employee_id or not employee_id.isdigit() or len(employee_id) < 6:
                    continue

                # Extraer commuting_allowance (通勤手当(非)) si existe
                commuting_value = 0
                if commuting_idx is not None and len(row_data) > commuting_idx:
                    commuting_value = self._to_number(row_data[commuting_idx])

                # Guardar el registro completo con todas las columnas
                full_record = {
                    "row_data": row_data,
                    "headers": headers,
                    "source_file": filename,
                    "commuting_idx": commuting_idx  # Índice dinámico de 通勤手当(非)
                }
                self.all_records.append(full_record)

                # Guardar en BD los campos principales para búsquedas
                db_record = {
                    "source_file": filename,
                    "employee_id": employee_id,
                    "name_roman": row_data[2] if len(row_data) > 2 else None,
                    "name_jp": str(row_data[3]).strip() if len(row_data) > 3 and row_data[3] else None,
                    "period": row_data[4] if len(row_data) > 4 else None,
                    "period_start": self._format_date(row_data[6]) if len(row_data) > 6 else None,
                    "period_end": self._format_date(row_data[7]) if len(row_data) > 7 else None,
                    "work_days": self._to_number(row_data[8]) if len(row_data) > 8 else 0,
                    "work_hours": self._to_number(row_data[12]) if len(row_data) > 12 else 0,
                    "overtime_hours": self._to_number(row_data[14]) if len(row_data) > 14 else 0,
                    "holiday_hours": 0,
                    "night_hours": self._to_number(row_data[16]) if len(row_data) > 16 else 0,
                    "hourly_rate": 0,
                    "base_pay": self._to_number(row_data[18]) if len(row_data) > 18 else 0,
                    "overtime_pay": self._to_number(row_data[19]) if len(row_data) > 19 else 0,
                    "night_pay": self._to_number(row_data[20]) if len(row_data) > 20 else 0,
                    "holiday_pay": self._to_number(row_data[21]) if len(row_data) > 21 else 0,
                    "commuting_allowance": commuting_value,
                    "total_pay": self._to_number(row_data[32]) if len(row_data) > 32 else 0,
                    "health_insurance": self._to_number(row_data[33]) if len(row_data) > 33 else 0,
                    "pension": self._to_number(row_data[34]) if len(row_data) > 34 else 0,
                    "employment_insurance": self._to_number(row_data[35]) if len(row_data) > 35 else 0,
                    "income_tax": self._to_number(row_data[38]) if len(row_data) > 38 else 0,
                    "resident_tax": self._to_number(row_data[37]) if len(row_data) > 37 else 0,
                    "deduction_total": self._to_number(row_data[48]) if len(row_data) > 48 else 0,
                    "net_pay": self._to_number(row_data[49]) if len(row_data) > 49 else 0,
                }
                
                save_payroll_record(db_record)
                records_count += 1
                self.records_saved += 1

            # Procesar hoja 請負 si existe (formato vertical para 請負社員)
            if "請負" in wb.sheetnames:
                print(f"   [Procesando] hoja 請負 (formato vertical)...")
                ws_ukeoi = wb["請負"]
                ukeoi_count = self.process_vertical_ukeoi_sheet(ws_ukeoi, filename)
                records_count += ukeoi_count
                self.records_saved += ukeoi_count
                print(f"   [INFO] Procesados {ukeoi_count} empleados 請負社員")

            wb.close()

            log_audit('PROCESS_FILE', 'processed_files', filename, None, None,
                      f"Procesados {records_count} registros")
            
            self.processed_files.append({
                "filename": filename,
                "records": records_count,
                "status": "success"
            })
            
            return {"status": "success", "records": records_count}
            
        except Exception as e:
            import traceback
            error_msg = f"Error procesando {filename}: {str(e)}"
            print(f"   [ERROR] {error_msg}")
            traceback.print_exc()
            self.errors.append(error_msg)
            log_audit('PROCESS_FILE_ERROR', 'processed_files', filename, None, None, str(e))
            
            self.processed_files.append({
                "filename": filename,
                "records": 0,
                "status": "error",
                "error": str(e)
            })
            
            return {"status": "error", "message": str(e)}

    def process_vertical_ukeoi_sheet(self, ws, filename: str) -> int:
        """
        Procesa hoja en formato vertical (請負社員)
        Cada empleado ocupa un bloque de ~14 columnas
        """
        records_count = 0

        # Detectar bloques de empleados buscando en todas las columnas
        # Buscar "給　料　支　払　明　細　書" o similar en fila 2
        employee_columns = []
        max_col = ws.max_column if ws.max_column else 1200

        for col in range(1, max_col + 1):
            val = ws.cell(row=2, column=col).value
            if val and '給' in str(val) and '明' in str(val) and '細' in str(val):
                employee_columns.append(col)

        print(f"   [INFO] Encontrados {len(employee_columns)} bloques de empleados 請負社員")

        # Procesar cada bloque de empleado
        for start_col in employee_columns:
            try:
                # Extraer employee_id (Fila 6, Col+8)
                employee_id = ws.cell(row=6, column=start_col + 8).value
                if not employee_id:
                    continue

                employee_id = str(employee_id).strip()
                if not employee_id.isdigit() or len(employee_id) < 6:
                    continue

                # Extraer período (Fila 5, Col+1)
                period = ws.cell(row=5, column=start_col + 1).value

                # Extraer nombre (Fila 8, Col+1) - formato "氏名 西岡　守"
                name_with_label = ws.cell(row=8, column=start_col + 1).value
                name_jp = None
                if name_with_label:
                    match = re.search(r'氏名\s*(.+)', str(name_with_label))
                    if match:
                        name_jp = match.group(1).strip()

                # Extraer datos de trabajo
                work_days = self._to_number(ws.cell(row=11, column=start_col + 4).value)
                work_hours = self._to_number(ws.cell(row=13, column=start_col + 2).value)
                overtime_hours = self._to_number(ws.cell(row=14, column=start_col + 2).value)
                night_hours = self._to_number(ws.cell(row=15, column=start_col + 2).value)

                # Extraer pagos
                base_pay = self._to_number(ws.cell(row=16, column=start_col + 2).value)
                overtime_pay = self._to_number(ws.cell(row=17, column=start_col + 2).value)
                night_pay = self._to_number(ws.cell(row=18, column=start_col + 2).value)
                commuting_allowance = self._to_number(ws.cell(row=20, column=start_col + 2).value)
                total_pay = self._to_number(ws.cell(row=30, column=start_col + 2).value)

                # Extraer deducciones
                health_insurance = self._to_number(ws.cell(row=31, column=start_col + 2).value)
                pension = self._to_number(ws.cell(row=32, column=start_col + 2).value)
                employment_insurance = self._to_number(ws.cell(row=33, column=start_col + 2).value)
                resident_tax = self._to_number(ws.cell(row=35, column=start_col + 2).value)
                income_tax = self._to_number(ws.cell(row=36, column=start_col + 2).value)
                deduction_total = self._to_number(ws.cell(row=46, column=start_col + 2).value)
                net_pay = self._to_number(ws.cell(row=47, column=start_col + 2).value)

                # Guardar en BD
                db_record = {
                    "source_file": filename,
                    "employee_id": employee_id,
                    "name_roman": None,
                    "name_jp": name_jp,
                    "period": period,
                    "period_start": None,
                    "period_end": None,
                    "work_days": work_days,
                    "work_hours": work_hours,
                    "overtime_hours": overtime_hours,
                    "holiday_hours": 0,
                    "night_hours": night_hours,
                    "hourly_rate": 0,
                    "base_pay": base_pay,
                    "overtime_pay": overtime_pay,
                    "night_pay": night_pay,
                    "holiday_pay": 0,
                    "total_pay": total_pay,
                    "health_insurance": health_insurance,
                    "pension": pension,
                    "employment_insurance": employment_insurance,
                    "income_tax": income_tax,
                    "resident_tax": resident_tax,
                    "deduction_total": deduction_total,
                    "net_pay": net_pay,
                }

                save_payroll_record(db_record)
                records_count += 1

            except Exception as e:
                print(f"   [ERROR] Procesando bloque col {start_col}: {e}")
                continue

        return records_count

    def _format_date(self, value):
        """Formatear fecha a string"""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return str(value) if value else None
    
    def _to_number(self, value):
        """Convertir a número"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return value
        try:
            return float(str(value).replace(",", "").replace("¥", ""))
        except:
            return 0
    
    def get_all_data(self) -> list:
        """Obtener todos los datos de la BD"""
        return get_all_payroll_records()
    
    def get_summary(self) -> dict:
        """Obtener resumen"""
        records = get_all_payroll_records()
        employees = get_all_employees()
        periods = get_periods()
        
        return {
            "total_records": len(records),
            "unique_employees": len(employees),
            "periods": periods,
            "processed_files": self.processed_files,
            "errors": self.errors,
            "records_saved_this_session": self.records_saved
        }
    
    def export_to_excel_all(self, output_path: str) -> str:
        """Exportar todos los datos a Excel ALL con TODAS las columnas"""
        if not self.all_records:
            # Si no hay registros en memoria, cargar de BD
            records = get_all_payroll_records()
            return self._export_from_db(output_path, records)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ALL"
        
        # Usar headers del primer registro
        headers = self.all_records[0]["headers"] if self.all_records else self.HEADERS_FULL
        
        # Estilo de headers
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Escribir datos
        for row_idx, record in enumerate(self.all_records, 2):
            row_data = record["row_data"]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                # Formato numérico para columnas de dinero
                if col in [19, 20, 21, 22, 23, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50]:
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0'
        
        # Ajustar anchos
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12
        
        wb.save(output_path)
        log_audit('EXPORT_ALL', None, None, None, None, f"Exportado a {output_path}")
        
        return output_path
    
    def export_by_month(self, output_path: str) -> str:
        """Exportar con hojas separadas por periodo (mes)"""
        if not self.all_records:
            records = get_all_payroll_records()
            return self._export_from_db_by_month(output_path, records)
        
        wb = Workbook()
        wb.remove(wb.active)
        
        # Agrupar por MES (extraer solo "2025年1月分" sin la fecha de pago)
        by_period = {}
        for record in self.all_records:
            row_data = record["row_data"]
            full_period = row_data[4] if len(row_data) > 4 and row_data[4] else "Unknown"
            
            # Extraer solo el mes: "2025年1月分(2月17日支給分)" -> "2025年1月分"
            match = re.match(r'(\d{4}年\d{1,2}月分)', str(full_period))
            if match:
                period = match.group(1)
            else:
                period = str(full_period)
            
            if period not in by_period:
                by_period[period] = []
            by_period[period].append(record)
        
        headers = self.all_records[0]["headers"] if self.all_records else self.HEADERS_FULL
        
        # Estilos
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        for period, period_records in sorted(by_period.items()):
            # Nombre de hoja seguro
            sheet_name = str(period)[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace("?", "").replace("[", "").replace("]", "")
            ws = wb.create_sheet(title=sheet_name)
            
            # Headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            # Datos
            for row_idx, record in enumerate(period_records, 2):
                row_data = record["row_data"]
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    if col in [19, 20, 21, 22, 23, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49, 50]:
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0'
        
        # Hoja ALL al final
        ws_all = wb.create_sheet(title="ALL")
        for col, header in enumerate(headers, 1):
            cell = ws_all.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, record in enumerate(self.all_records, 2):
            row_data = record["row_data"]
            for col, value in enumerate(row_data, 1):
                cell = ws_all.cell(row=row_idx, column=col, value=value)
                if isinstance(value, (int, float)) and col > 18:
                    cell.number_format = '#,##0'
        
        wb.save(output_path)
        log_audit('EXPORT_BY_MONTH', None, None, None, None, f"Exportado a {output_path}")
        
        return output_path
    
    def _export_from_db(self, output_path: str, records: list) -> str:
        """Exportar desde BD cuando no hay registros en memoria"""
        wb = Workbook()
        ws = wb.active
        ws.title = "ALL"
        
        headers = self.HEADERS_FULL[:50]  # Solo los campos que tenemos en BD
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row_idx, record in enumerate(records, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=record.get('employee_id'))
            ws.cell(row=row_idx, column=3, value=record.get('name_roman'))
            ws.cell(row=row_idx, column=4, value=record.get('name_jp'))
            ws.cell(row=row_idx, column=5, value=record.get('period'))
            ws.cell(row=row_idx, column=7, value=record.get('period_start'))
            ws.cell(row=row_idx, column=8, value=record.get('period_end'))
            ws.cell(row=row_idx, column=9, value=record.get('work_days'))
            ws.cell(row=row_idx, column=13, value=record.get('work_hours'))
            ws.cell(row=row_idx, column=15, value=record.get('overtime_hours'))
            ws.cell(row=row_idx, column=17, value=record.get('night_hours'))
            ws.cell(row=row_idx, column=19, value=record.get('base_pay'))
            ws.cell(row=row_idx, column=20, value=record.get('overtime_pay'))
            ws.cell(row=row_idx, column=21, value=record.get('night_pay'))
            ws.cell(row=row_idx, column=22, value=record.get('holiday_pay'))
            ws.cell(row=row_idx, column=33, value=record.get('total_pay'))
            ws.cell(row=row_idx, column=34, value=record.get('health_insurance'))
            ws.cell(row=row_idx, column=35, value=record.get('pension'))
            ws.cell(row=row_idx, column=36, value=record.get('employment_insurance'))
            ws.cell(row=row_idx, column=38, value=record.get('resident_tax'))
            ws.cell(row=row_idx, column=39, value=record.get('income_tax'))
            ws.cell(row=row_idx, column=49, value=record.get('deduction_total'))
            ws.cell(row=row_idx, column=50, value=record.get('net_pay'))
        
        wb.save(output_path)
        return output_path
    
    def _export_from_db_by_month(self, output_path: str, records: list) -> str:
        """Exportar por mes desde BD"""
        wb = Workbook()
        wb.remove(wb.active)
        
        by_period = {}
        for record in records:
            full_period = record.get("period", "Unknown")
            
            # Extraer solo el mes: "2025年1月分(2月17日支給分)" -> "2025年1月分"
            match = re.match(r'(\d{4}年\d{1,2}月分)', str(full_period))
            if match:
                period = match.group(1)
            else:
                period = str(full_period)
            
            if period not in by_period:
                by_period[period] = []
            by_period[period].append(record)
        
        headers = self.HEADERS_FULL[:50]
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        
        for period, period_records in sorted(by_period.items()):
            sheet_name = str(period)[:31].replace("/", "-").replace("\\", "-")
            ws = wb.create_sheet(title=sheet_name)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            for row_idx, record in enumerate(period_records, 2):
                ws.cell(row=row_idx, column=2, value=record.get('employee_id'))
                ws.cell(row=row_idx, column=3, value=record.get('name_roman'))
                ws.cell(row=row_idx, column=4, value=record.get('name_jp'))
                ws.cell(row=row_idx, column=5, value=record.get('period'))
                ws.cell(row=row_idx, column=33, value=record.get('total_pay'))
                ws.cell(row=row_idx, column=50, value=record.get('net_pay'))
        
        wb.save(output_path)
        return output_path
    
    def export_chingin_by_employee(self, output_folder: str) -> list:
        """Exportar 賃金台帳 individual por empleado"""
        os.makedirs(output_folder, exist_ok=True)
        
        employees = get_all_employees()
        generated_files = []
        
        for emp in employees:
            emp_id = emp['employee_id']
            records = get_payroll_by_employee(emp_id)
            
            if not records:
                continue
            
            wb = Workbook()
            ws = wb.active
            ws.title = "賃金台帳"
            
            # Título
            ws['B2'] = f"賃金台帳 - {datetime.now().year}年"
            ws['B2'].font = Font(bold=True, size=16)
            
            # Info empleado
            ws['B4'] = "従業員番号"
            ws['C4'] = emp_id
            ws['B5'] = "氏名"
            ws['C5'] = emp.get('name_jp', records[0].get('name_jp', ''))
            ws['B6'] = "氏名ローマ字"
            ws['C6'] = emp.get('name_roman', records[0].get('name_roman', ''))
            
            # Headers de meses
            row = 9
            headers = ["項目", "1月", "2月", "3月", "4月", "5月", "6月",
                      "7月", "8月", "9月", "10月", "11月", "12月", "合計"]
            
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="E0E0E0")
            
            # Items
            items = [
                ("出勤日数", "work_days"),
                ("実働時間", "work_hours"),
                ("残業時間", "overtime_hours"),
                ("深夜時間", "night_hours"),
                ("基本給", "base_pay"),
                ("残業手当", "overtime_pay"),
                ("深夜手当", "night_pay"),
                ("総支給額", "total_pay"),
                ("健康保険", "health_insurance"),
                ("厚生年金", "pension"),
                ("雇用保険", "employment_insurance"),
                ("所得税", "income_tax"),
                ("控除合計", "deduction_total"),
                ("差引支給額", "net_pay")
            ]
            
            # Mapear por mes
            by_month = {}
            for rec in records:
                period = rec.get("period", "")
                match = re.search(r'(\d+)月', period)
                if match:
                    month = int(match.group(1))
                    by_month[month] = rec
            
            for item_idx, (item_name, field) in enumerate(items):
                r = row + item_idx + 1
                ws.cell(row=r, column=1, value=item_name)
                
                total = 0
                for month in range(1, 13):
                    col = month + 1
                    if month in by_month:
                        value = by_month[month].get(field, 0) or 0
                        ws.cell(row=r, column=col, value=value)
                        if isinstance(value, (int, float)):
                            total += value
                            ws.cell(row=r, column=col).number_format = '#,##0'
                
                ws.cell(row=r, column=14, value=total)
                ws.cell(row=r, column=14).number_format = '#,##0'
                ws.cell(row=r, column=14).font = Font(bold=True)
            
            # Guardar
            safe_name = emp_id.replace("/", "_").replace("\\", "_")
            filepath = os.path.join(output_folder, f"賃金台帳_{safe_name}.xlsx")
            wb.save(filepath)
            generated_files.append(filepath)
        
        log_audit('EXPORT_CHINGIN', None, None, None, None,
                  f"Exportados {len(generated_files)} archivos 賃金台帳")
        
        return generated_files
    
    def clear(self):
        """Limpiar datos de sesión"""
        self.processed_files = []
        self.errors = []
        self.records_saved = 0
        self.all_records = []
    
    def generate_chingin_print(self, employee_id: str, year: int = None, output_path: str = None) -> dict:
        """
        Genera 賃金台帳 para un empleado en formato Print (como la hoja Print del archivo XP)
        Similar a la hoja Print que usa XLOOKUP para buscar datos por ID
        
        Args:
            employee_id: ID del empleado (従業員番号)
            year: Año a generar (default: año actual)
            output_path: Ruta de salida (opcional)
        
        Returns:
            dict con info del empleado y path del archivo generado
        """
        from database import get_payroll_by_employee, get_all_employees, get_employee_master
        
        if year is None:
            year = datetime.now().year
        
        # Buscar datos del empleado
        records = get_payroll_by_employee(employee_id)
        
        if not records:
            return {"error": f"No se encontraron datos para el empleado {employee_id}"}
        
        # Info del empleado
        emp_info = {
            "employee_id": employee_id,
            "name_jp": records[0].get('name_jp', ''),
            "name_roman": records[0].get('name_roman', ''),
        }
        
        # Buscar datos del maestro de empleados (派遣社員/請負社員)
        master_data = get_employee_master(employee_id)
        
        # Buscar datos adicionales del empleado (派遣先, 性別, etc.) del primer registro
        # También podemos buscar en all_records si hay datos en memoria
        dispatch = ""
        if master_data:
            dispatch = master_data.get('dispatch_company', '') or master_data.get('job_type', '')
        else:
            for rec in self.all_records:
                if str(rec["row_data"][1]) == str(employee_id):
                    dispatch = rec["row_data"][5] if len(rec["row_data"]) > 5 else ""
                    break
        
        # Organizar por mes
        by_month = {}
        for rec in records:
            period = rec.get("period", "")
            match = re.search(r'(\d+)月', str(period))
            if match:
                month = int(match.group(1))
                by_month[month] = rec
        
        # También buscar en all_records para datos completos
        for rec in self.all_records:
            if str(rec["row_data"][1]) == str(employee_id):
                period = rec["row_data"][4] if len(rec["row_data"]) > 4 else ""
                match = re.search(r'(\d+)月', str(period))
                if match:
                    month = int(match.group(1))
                    # Almacenar datos completos incluyendo commuting_idx dinámico
                    by_month[month] = {
                        "full_data": rec["row_data"],
                        "commuting_idx": rec.get("commuting_idx"),  # Índice dinámico de 通勤手当(非)
                        **by_month.get(month, {})
                    }
        
        # Crear workbook con formato Print
        wb = Workbook()
        ws = wb.active
        ws.title = "賃金台帳"
        
        # Configurar ancho de columnas
        ws.column_dimensions['A'].width = 3
        ws.column_dimensions['B'].width = 18
        for col in 'CDEFGHIJKLMNO':
            ws.column_dimensions[col].width = 10
        ws.column_dimensions['O'].width = 12
        
        # Estilos
        header_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=14)
        money_format = '#,##0'
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill("solid", fgColor="E0E0E0")
        
        # Obtener datos del empleado del maestro (入社日, 性別)
        hire_date = ""
        gender = ""
        emp_name = emp_info.get('name_jp', '') or emp_info.get('name_roman', '')
        
        if master_data:
            # Usar datos del maestro sincronizado
            hire_date = master_data.get('hire_date', '') or ""
            gender = master_data.get('gender', '') or ""
            emp_name = master_data.get('name', '') or emp_name
        else:
            # Fallback: buscar en registros de nómina
            for rec in self.all_records:
                if str(rec["row_data"][1]) == str(employee_id):
                    headers = rec.get("headers", [])
                    row_data = rec["row_data"]
                    # Buscar 入社日 y 性別 en headers
                    for idx, h in enumerate(headers):
                        if h and idx < len(row_data):
                            h_str = str(h)
                            if '入社' in h_str and not hire_date:
                                hire_date = row_data[idx] if row_data[idx] else ""
                            if '性別' in h_str and not gender:
                                gender = row_data[idx] if row_data[idx] else ""
                    break
        
        # === ENCABEZADO ===
        ws['B1'] = "入社日"
        ws['C1'] = hire_date  # Fecha de entrada del empleado
        if hire_date and hasattr(hire_date, 'strftime'):
            ws['C1'] = hire_date.strftime("%Y/%m/%d")
        
        ws['B2'] = "従業員番号"
        ws['C2'] = "氏      名"
        ws['G2'] = "性別"
        ws['H2'] = year
        ws['H2'].font = title_font
        ws['J2'] = "賃金台帳"
        ws['J2'].font = title_font
        
        ws['B3'] = employee_id
        ws['B3'].font = Font(bold=True, size=12)
        ws['C3'] = emp_name  # Nombre del empleado (del maestro o de nómina)
        ws['C3'].font = Font(bold=True, size=12)
        ws['G3'] = gender  # Sexo del empleado
        
        ws['B4'] = "派遣先＜＞所属先"
        ws['C4'] = dispatch
        
        # === MESES (Fila 6) - Formato "1月分", "2月分", etc. ===
        for month in range(1, 13):
            col = month + 2  # C=1, D=2, ... N=12
            cell = ws.cell(row=6, column=col, value=f"{month}月分")
            cell.alignment = Alignment(horizontal="center")
            cell.font = header_font
        ws.cell(row=6, column=15, value="合  計")
        ws.cell(row=6, column=15).font = header_font
        ws.cell(row=6, column=15).alignment = Alignment(horizontal="center")
        
        # === DATOS POR FILA ===
        # Estructura de filas como en la hoja Print
        rows_config = [
            (7, "支給分", None),  # Mostrar periodo
            (8, "賃金計算期間", None),  # Mostrar rango de fechas
            (9, "出勤日数", "work_days"),
            (10, "休日出勤日数", None),
            (11, "欠勤日数", "absence_days"),
            (12, "有休日数", "paid_leave_days"),
            (13, "特別休暇日数", None),
            (14, "実働時間", "work_hours"),  # Formato especial H:MM
            (15, "残業時間数", "overtime_hours"),  # Formato especial H:MM
            (16, "休日労働時間数", None),
            (17, "深夜労働時間数", "night_hours"),  # Formato especial H:MM
            (18, "基本給 (月給)", None),
            (19, "基本給 (日給)", None),
            (20, "基本給 (時給)", "base_pay"),
            (21, "役員報酬", None),
            (22, "職務給", None),
            (23, "役付手当", None),
            (24, "家族手当", None),
            (25, "住宅手当", None),
            (26, "資格手当", None),
            (27, "営業外勤手当", None),
            (28, "その他手当１", "sum_allowances_1_8"),  # SUMA de columnas X(1) a AE(8)
            (29, "その他手当２", None),
            (30, "その他手当３", None),
            (31, "その他手当４", None),
            (32, "その他手当５", None),
            (33, "その他手当１(前月)", None),
            (34, "その他手当２(前月)", None),
            (35, "その他手当３(前月)", None),
            (36, "その他手当４(前月)", None),
            (37, "その他手当５(前月)", None),
            (38, "課税通勤費", None),
            (39, "非課税通勤費", "commuting_allowance"),
            (40, "普通残業手当", "overtime_pay"),
            (41, "深夜残業手当", "night_pay"),
            (42, "休日勤務手当", "holiday_pay"),
            (43, "欠勤遅刻早退控除", None),
            (44, "欠勤遅刻早退控除(前月)", None),
            (45, "前月修正１", None),
            (46, "前月修正２", None),
            (47, "前月修正３", None),
            (48, "前月修正４", None),
            (49, "前月修正５", None),
            (50, "前々月修正１", None),
            (51, "前々月修正２", None),
            (52, "前々月修正３", None),
            (53, "前々月修正４", None),
            (54, "前々月修正５", None),
            (55, "休業補償費", None),
            (56, "課税現物給与", None),
            (57, "非課税現物給与", None),
            (58, "課税昇給差額", None),
            (59, "非課税昇給差額", None),
            (60, "賞与", None),
            (61, "現物賞与", None),
            (62, "役員賞与", None),
            (63, "課税支給合計", None),  # Calculado
            (64, "非課税支給合計", None),  # Calculado
            (65, "支給合計", "total_pay"),
            (66, "健康保険料", "health_insurance"),
            (67, "介護保険料", None),
            (68, "厚生年金保険料", "pension"),
            (69, "厚生年金基金保険料", None),
            (70, "社保料調整", None),
            (71, "雇用保険料", "employment_insurance"),
            (72, "所得税", "income_tax"),
            (73, "住民税", "resident_tax"),
            (74, "財形貯蓄", None),
            (75, "組合費", None),
            (76, "その他", "sum_deductions_1_8"),  # SUMA de columnas AN(控除1) a AU(控除8)
            (77, "控除合計", "deduction_total"),
            (78, "年末調整還付", "nencho_refund"),  # 控除9 cuando es negativo (devolución)
            (79, "年末調整徴収", "nencho_collect"),  # 控除9 cuando es positivo (cobro)
            (80, "差引支給額", "net_pay"),
        ]
        
        # Mapeo de campos de all_records (indices de columna - 0-based)
        # Columnas: X(1)=23, Y(2)=24, Z(3)=25, AA(4)=26, AB(5)=27, AC(6)=28, AD(7)=29, AE(8)=30
        # Columnas: AN(控除1)=39, AO(控除2)=40, AP(控除3)=41, AQ(控除4)=42, AR(控除5)=43, AS(控除6)=44, AT(控除7)=45, AU(控除8)=46
        # Columna:  AV(控除9)=47
        field_to_idx = {
            "work_days": 8,
            "absence_days": 9,
            "paid_leave_days": 10,
            "work_hours": 12,  # Necesita combinar con 13 (minutos)
            "work_minutes": 13,
            "overtime_hours": 14,  # Necesita combinar con 15 (minutos)
            "overtime_minutes": 15,
            "night_hours": 16,  # Necesita combinar con 17 (minutos)
            "night_minutes": 17,
            "base_pay": 18,
            "overtime_pay": 19,
            "night_pay": 20,
            "holiday_pay": 21,
            "paid_leave_pay": 22,
            "commuting_allowance": 50,
            "total_pay": 32,
            "health_insurance": 33,
            "pension": 34,
            "employment_insurance": 35,
            "income_tax": 38,
            "resident_tax": 37,
            "deduction_total": 48,
            "net_pay": 49,
            # Campos calculados especiales
            "sum_allowances_1_8": [23, 24, 25, 26, 27, 28, 29, 30],  # X(1) a AE(8)
            "sum_deductions_1_8": [39, 40, 41, 42, 43, 44, 45, 46],  # AN(控除1) a AU(控除8)
            "nencho_adjustment": 47,  # AV(控除9) - 年調過不足
        }
        
        # Escribir datos
        for row_num, label, field in rows_config:
            # Label
            cell = ws.cell(row=row_num, column=2, value=label)
            cell.border = border_thin
            
            total = 0
            has_data = False
            
            for month in range(1, 13):
                col = month + 2
                cell = ws.cell(row=row_num, column=col)
                cell.border = border_thin
                
                if month in by_month:
                    rec = by_month[month]
                    value = None
                    
                    # Fila 7: Mostrar periodo completo
                    if row_num == 7:
                        if "full_data" in rec:
                            value = rec["full_data"][4] if len(rec["full_data"]) > 4 else None
                        else:
                            value = rec.get("period", "")
                        if value:
                            # Extraer solo la parte de fecha de pago
                            match = re.search(r'\(([^)]+)\)', str(value))
                            if match:
                                value = match.group(1)
                    
                    # Fila 8: Periodo de cálculo
                    elif row_num == 8:
                        if "full_data" in rec:
                            start = rec["full_data"][6] if len(rec["full_data"]) > 6 else None
                            end = rec["full_data"][7] if len(rec["full_data"]) > 7 else None
                            if start and end:
                                start_str = start.strftime("%m/%d") if hasattr(start, 'strftime') else str(start)[:5]
                                end_str = end.strftime("%m/%d") if hasattr(end, 'strftime') else str(end)[:5]
                                value = f"{start_str}～{end_str}"
                        else:
                            start = rec.get("period_start", "")
                            end = rec.get("period_end", "")
                            if start and end:
                                value = f"{str(start)[5:10]}～{str(end)[5:10]}"
                    
                    # Horas con formato H:MM
                    elif field in ["work_hours", "overtime_hours", "night_hours"]:
                        if "full_data" in rec:
                            hours_idx = field_to_idx.get(field, 0)
                            mins_idx = hours_idx + 1
                            hours = rec["full_data"][hours_idx] if len(rec["full_data"]) > hours_idx else 0
                            mins = rec["full_data"][mins_idx] if len(rec["full_data"]) > mins_idx else 0
                            hours = hours or 0
                            mins = mins or 0
                            if hours or mins:
                                value = f"{int(hours)}:{int(mins):02d}"
                                has_data = True
                        else:
                            # Convertir horas decimales a H:MM
                            # Ej: 153.4 = 153 horas + 0.4*60 = 24 minutos = 153:24
                            decimal_hours = rec.get(field, 0) or 0
                            if decimal_hours:
                                total_hours = int(decimal_hours)
                                total_mins = int(round((decimal_hours - total_hours) * 60))
                                value = f"{total_hours}:{total_mins:02d}"
                                has_data = True
                    
                    # CAMPO ESPECIAL: その他手当１ = SUMA de columnas X(1) a AE(8)
                    # EXCLUIR 通勤手当(非) que se muestra por separado en fila 39
                    elif field == "sum_allowances_1_8":
                        if "full_data" in rec:
                            indices = field_to_idx["sum_allowances_1_8"]
                            sum_val = 0
                            for idx in indices:
                                v = rec["full_data"][idx] if len(rec["full_data"]) > idx else 0
                                if v and isinstance(v, (int, float)):
                                    sum_val += v
                            # Restar commuting_allowance (通勤手当(非)) del total
                            commuting = rec.get("commuting_allowance", 0)
                            if commuting and isinstance(commuting, (int, float)):
                                sum_val -= commuting
                            if sum_val != 0:
                                value = sum_val
                                total += sum_val
                                has_data = True
                                cell.number_format = money_format
                    
                    # CAMPO ESPECIAL: その他 = SUMA de columnas AN(控除1) a AU(控除8)
                    elif field == "sum_deductions_1_8":
                        if "full_data" in rec:
                            indices = field_to_idx["sum_deductions_1_8"]
                            sum_val = 0
                            for idx in indices:
                                v = rec["full_data"][idx] if len(rec["full_data"]) > idx else 0
                                if v and isinstance(v, (int, float)):
                                    sum_val += v
                            if sum_val != 0:
                                value = sum_val
                                total += sum_val
                                has_data = True
                                cell.number_format = money_format
                    
                    # CAMPO ESPECIAL: 年末調整還付 (devolución) - cuando 控除9 es negativo
                    elif field == "nencho_refund":
                        if "full_data" in rec:
                            idx = field_to_idx["nencho_adjustment"]
                            v = rec["full_data"][idx] if len(rec["full_data"]) > idx else 0
                            if v and isinstance(v, (int, float)) and v < 0:
                                value = abs(v)  # Mostrar como positivo
                                total += abs(v)
                                has_data = True
                                cell.number_format = money_format
                    
                    # CAMPO ESPECIAL: 年末調整徴収 (cobro) - cuando 控除9 es positivo
                    elif field == "nencho_collect":
                        if "full_data" in rec:
                            idx = field_to_idx["nencho_adjustment"]
                            v = rec["full_data"][idx] if len(rec["full_data"]) > idx else 0
                            if v and isinstance(v, (int, float)) and v > 0:
                                value = v
                                total += v
                                has_data = True
                                cell.number_format = money_format
                    
                    # Campos numéricos normales
                    elif field and field in field_to_idx and not isinstance(field_to_idx.get(field), list):
                        if "full_data" in rec:
                            idx = field_to_idx[field]
                            value = rec["full_data"][idx] if len(rec["full_data"]) > idx else 0
                        else:
                            value = rec.get(field, 0)
                        
                        if value and isinstance(value, (int, float)) and value != 0:
                            total += value
                            has_data = True
                            cell.number_format = money_format
                    
                    elif field:
                        value = rec.get(field, None)
                        if value and isinstance(value, (int, float)) and value != 0:
                            total += value
                            has_data = True
                            cell.number_format = money_format
                    
                    if value:
                        cell.value = value
                        cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "center")
            
            # Columna Total
            total_cell = ws.cell(row=row_num, column=15)
            total_cell.border = border_thin
            if has_data and total:
                total_cell.value = total
                total_cell.number_format = money_format
                total_cell.font = Font(bold=True)
            total_cell.alignment = Alignment(horizontal="right")
        
        # Aplicar bordes al encabezado de meses
        for col in range(2, 16):
            ws.cell(row=6, column=col).border = border_thin
            ws.cell(row=6, column=col).fill = header_fill
        
        # Guardar
        if output_path is None:
            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "outputs")
            os.makedirs(output_dir, exist_ok=True)
            output_path = os.path.join(output_dir, f"賃金台帳_{employee_id}_{year}.xlsx")
        
        wb.save(output_path)
        
        log_audit('GENERATE_CHINGIN_PRINT', 'employees', employee_id, None, None,
                  f"Generado 賃金台帳 formato Print para {employee_id}")
        
        return {
            "success": True,
            "employee_id": employee_id,
            "name": emp_info.get('name_jp', '') or emp_info.get('name_roman', ''),
            "year": year,
            "months_found": list(by_month.keys()),
            "output_path": output_path
        }
    
    def search_employee(self, employee_id: str) -> dict:
        """Buscar empleado y retornar su informacion"""
        from database import get_payroll_by_employee, get_employee_master
        
        records = get_payroll_by_employee(employee_id)
        
        # Buscar datos maestros del empleado
        master_data = get_employee_master(employee_id)
        name_jp = ""
        name_roman = ""
        dispatch = ""
        
        if master_data:
            # name es el nombre romano, name_kana es el nombre en katakana
            name_roman = master_data.get('name', '') or ""
            name_kana = master_data.get('name_kana', '') or ""
            # Mostrar ambos nombres: romano y kana
            name_jp = name_kana if name_kana else name_roman
            dispatch = master_data.get('dispatch_company', '') or master_data.get('job_type', '') or ""
        
        if not records:
            # Buscar en all_records
            for rec in self.all_records:
                if str(rec["row_data"][1]) == str(employee_id):
                    return {
                        "found": True,
                        "employee_id": employee_id,
                        "name_jp": name_jp or (rec["row_data"][3] if len(rec["row_data"]) > 3 else ""),
                        "name_roman": name_roman or (rec["row_data"][2] if len(rec["row_data"]) > 2 else ""),
                        "dispatch": dispatch or (rec["row_data"][5] if len(rec["row_data"]) > 5 else ""),
                        "records": 1
                    }
            
            # Si hay datos maestros pero no registros
            if master_data:
                return {
                    "found": True,
                    "employee_id": employee_id,
                    "name_jp": name_jp,
                    "name_roman": name_roman,
                    "dispatch": dispatch,
                    "records": 0,
                    "periods": []
                }
            
            return {"found": False, "employee_id": employee_id}
        
        # Contar periodos únicos
        periods = set()
        for rec in records:
            if rec.get("period"):
                periods.add(rec["period"])
        
        return {
            "found": True,
            "employee_id": employee_id,
            "name_jp": name_jp or records[0].get('name_jp', ''),
            "name_roman": name_roman or records[0].get('name_roman', ''),
            "dispatch": dispatch,
            "records": len(records),
            "periods": list(periods)
        }

    def generate_chingin_format_b(self, employee_id: str, year: int = None, output_path: str = None) -> dict:
        """
        Generar Chingin台帳 usando Template B (formato horizontal 12 meses)

        Args:
            employee_id: ID del empleado
            year: Ano (default: ano actual)
            output_path: Ruta de salida (opcional)

        Returns:
            dict con status, message y file_path
        """
        try:
            if year is None:
                year = datetime.now().year

            print(f"\n[INFO] Generando Chingin Format B para empleado {employee_id}, ano {year}")

            # Cargar template
            template_path = os.path.join("templates", "template_format_b.xlsx")
            if not os.path.exists(template_path):
                return {
                    "status": "error",
                    "message": f"Template B no encontrado: {template_path}"
                }

            wb = load_workbook(template_path)
            ws = wb.active

            # Obtener datos del empleado para todos los meses del ano
            from database import get_payroll_by_employee_year
            records = get_payroll_by_employee_year(employee_id, year)

            if not records:
                wb.close()
                return {
                    "status": "error",
                    "message": f"No hay datos para empleado {employee_id} en {year}"
                }

            # Obtener informacion del empleado
            first_record = records[0]
            employee_name = first_record.get('name_jp', '')
            birth_date = first_record.get('birth_date', '')
            hire_date = first_record.get('hire_date', '')
            department = first_record.get('department', '')
            gender = first_record.get('gender', '')

            print(f"[INFO] Empleado: {employee_name}")
            print(f"[INFO] Registros encontrados: {len(records)}")

            # LLENAR DATOS DEL TEMPLATE

            # Row 1: Titulo con ano
            ws['A1'] = f"  {year}年　　賃　金　台　帳"

            # Row 4: Informacion del empleado
            # E4-G4: Birth date
            if birth_date:
                try:
                    bd = datetime.strptime(birth_date, '%Y-%m-%d')
                    ws['E4'] = f"{bd.year}年"
                    ws['F4'] = f"{bd.month}月"
                    ws['G4'] = f"{bd.day}日"
                except:
                    ws['E4'] = birth_date

            # H4-J4: Hire date
            if hire_date:
                try:
                    hd = datetime.strptime(hire_date, '%Y-%m-%d')
                    ws['H4'] = f"{hd.year}年"
                    ws['I4'] = f"{hd.month}月"
                    ws['J4'] = f"{hd.day}日"
                except:
                    ws['H4'] = hire_date

            # K4: Department
            ws['K4'] = department

            # M4: Name
            ws['M4'] = employee_name

            # P4: Gender
            ws['P4'] = gender

            # Organizar records por mes (1-12)
            records_by_month = {}
            for rec in records:
                period = rec.get('period', '')
                # Extraer mes del periodo (formato: YYYY-MM)
                if '-' in period:
                    month = int(period.split('-')[1])
                    records_by_month[month] = rec

            # Mapeo de campos DB -> filas del template
            field_mapping = {
                7: 'work_days',          # 労働日数
                8: 'work_hours',         # 労働時間数
                9: 'overtime_hours',     # 時間外労働
                10: 'holiday_hours',     # 休日労働
                11: 'night_hours',       # 深夜労働
                13: 'base_pay',          # 基本給
                18: 'overtime_pay',      # 時間外手当
                19: 'holiday_pay',       # 休日労働手当
                20: 'night_pay',         # 深夜勤務手当
                22: 'commuting_allowance', # 通勤手当(非課税)
                27: 'health_insurance',  # 健康保険
                28: 'care_insurance',    # 介護保険
                29: 'pension_insurance', # 厚生年金
                30: 'employment_insurance', # 雇用保険
                33: 'income_tax',        # 所得税
                34: 'resident_tax',      # 住民税
            }

            # Llenar datos por mes (columnas B-M = columnas 2-13)
            for month in range(1, 13):
                col = month + 1  # B=2, C=3, ..., M=13

                if month in records_by_month:
                    rec = records_by_month[month]

                    # Llenar cada campo segun el mapeo
                    for row, field in field_mapping.items():
                        value = rec.get(field, 0)
                        if value and value != 0:
                            cell = ws.cell(row=row, column=col)
                            cell.value = value

                            # Aplicar formato segun tipo de dato
                            if field in ['work_days']:
                                cell.number_format = '0"日"'
                            elif field in ['work_hours', 'overtime_hours', 'holiday_hours', 'night_hours']:
                                cell.number_format = '0"時間"'
                            else:
                                cell.number_format = '#,##0'

            # Calcular totales en columna P (columna 16)
            for row, field in field_mapping.items():
                total = 0
                for month in range(1, 13):
                    if month in records_by_month:
                        value = records_by_month[month].get(field, 0)
                        if value:
                            total += value

                if total != 0:
                    cell = ws.cell(row=row, column=16)  # Column P
                    cell.value = total

                    # Aplicar formato
                    if field in ['work_days']:
                        cell.number_format = '0"日"'
                    elif field in ['work_hours', 'overtime_hours', 'holiday_hours', 'night_hours']:
                        cell.number_format = '0"時間"'
                    else:
                        cell.number_format = '#,##0'

            # Calcular filas de totales especiales
            for month in range(1, 13):
                col = month + 1

                if month in records_by_month:
                    rec = records_by_month[month]

                    # Row 24: 課税合計 (Taxable total)
                    taxable_total = (rec.get('base_pay', 0) +
                                    rec.get('overtime_pay', 0) +
                                    rec.get('holiday_pay', 0) +
                                    rec.get('night_pay', 0))
                    if taxable_total > 0:
                        ws.cell(row=24, column=col).value = taxable_total
                        ws.cell(row=24, column=col).number_format = '#,##0'

                    # Row 25: 非課税合計 (Non-taxable total)
                    non_taxable = rec.get('commuting_allowance', 0)
                    if non_taxable > 0:
                        ws.cell(row=25, column=col).value = non_taxable
                        ws.cell(row=25, column=col).number_format = '#,##0'

                    # Row 26: 総支給合計 (Gross total)
                    gross_total = taxable_total + non_taxable
                    if gross_total > 0:
                        ws.cell(row=26, column=col).value = gross_total
                        ws.cell(row=26, column=col).number_format = '#,##0'

                    # Row 31: 社会保険合計 (Social insurance total)
                    social_ins_total = (rec.get('health_insurance', 0) +
                                       rec.get('care_insurance', 0) +
                                       rec.get('pension_insurance', 0) +
                                       rec.get('employment_insurance', 0))
                    if social_ins_total > 0:
                        ws.cell(row=31, column=col).value = social_ins_total
                        ws.cell(row=31, column=col).number_format = '#,##0'

                    # Row 32: 課税対象額 (Taxable amount)
                    taxable_amount = taxable_total
                    if taxable_amount > 0:
                        ws.cell(row=32, column=col).value = taxable_amount
                        ws.cell(row=32, column=col).number_format = '#,##0'

                    # Row 39: 控除合計 (Total deductions)
                    total_deductions = (social_ins_total +
                                       rec.get('income_tax', 0) +
                                       rec.get('resident_tax', 0))
                    if total_deductions > 0:
                        ws.cell(row=39, column=col).value = total_deductions
                        ws.cell(row=39, column=col).number_format = '#,##0'

                    # Row 40: 差引支給額 (Net pay)
                    net_pay = gross_total - total_deductions
                    if net_pay > 0:
                        ws.cell(row=40, column=col).value = net_pay
                        ws.cell(row=40, column=col).number_format = '#,##0'

            # Calcular totales anuales en columna P (column 16)
            # Row 24: Total taxable
            total_taxable = sum(records_by_month[m].get('base_pay', 0) +
                               records_by_month[m].get('overtime_pay', 0) +
                               records_by_month[m].get('holiday_pay', 0) +
                               records_by_month[m].get('night_pay', 0)
                               for m in records_by_month)
            if total_taxable > 0:
                ws.cell(row=24, column=16).value = total_taxable
                ws.cell(row=24, column=16).number_format = '#,##0'

            # Row 25: Total non-taxable
            total_non_taxable = sum(records_by_month[m].get('commuting_allowance', 0)
                                   for m in records_by_month)
            if total_non_taxable > 0:
                ws.cell(row=25, column=16).value = total_non_taxable
                ws.cell(row=25, column=16).number_format = '#,##0'

            # Row 26: Gross total
            total_gross = total_taxable + total_non_taxable
            if total_gross > 0:
                ws.cell(row=26, column=16).value = total_gross
                ws.cell(row=26, column=16).number_format = '#,##0'

            # Row 31: Total social insurance
            total_social = sum(records_by_month[m].get('health_insurance', 0) +
                              records_by_month[m].get('care_insurance', 0) +
                              records_by_month[m].get('pension_insurance', 0) +
                              records_by_month[m].get('employment_insurance', 0)
                              for m in records_by_month)
            if total_social > 0:
                ws.cell(row=31, column=16).value = total_social
                ws.cell(row=31, column=16).number_format = '#,##0'

            # Row 39: Total deductions
            total_deduct = (total_social +
                           sum(records_by_month[m].get('income_tax', 0) for m in records_by_month) +
                           sum(records_by_month[m].get('resident_tax', 0) for m in records_by_month))
            if total_deduct > 0:
                ws.cell(row=39, column=16).value = total_deduct
                ws.cell(row=39, column=16).number_format = '#,##0'

            # Row 40: Net pay
            total_net = total_gross - total_deduct
            if total_net > 0:
                ws.cell(row=40, column=16).value = total_net
                ws.cell(row=40, column=16).number_format = '#,##0'

            # Guardar archivo
            if output_path is None:
                os.makedirs("outputs", exist_ok=True)
                safe_name = employee_name.replace('/', '_').replace('\\', '_')
                output_path = f"outputs/賃金台帳_{safe_name}_{year}_FormatB.xlsx"

            wb.save(output_path)
            wb.close()

            print(f"[OK] Archivo generado: {output_path}")

            return {
                "status": "success",
                "message": "Chingin Format B generado exitosamente",
                "file_path": output_path,
                "employee_id": employee_id,
                "employee_name": employee_name,
                "year": year,
                "records": len(records)
            }

        except Exception as e:
            print(f"[ERROR] generate_chingin_format_b: {e}")
            import traceback
            traceback.print_exc()
            return {
                "status": "error",
                "message": str(e)
            }

    def generate_chingin_format_c(self, employee_id: str, year: int = None, output_path: str = None) -> dict:
        """
        Generar Chingin台帳 usando Template C (formato horizontal 12 meses simplificado)

        Args:
            employee_id: ID del empleado
            year: Ano (default: ano actual)
            output_path: Ruta de salida (opcional)

        Returns:
            dict con status, message y file_path
        """
        try:
            if year is None:
                year = datetime.now().year

            print(f"\n[INFO] Generando Chingin Format C para empleado {employee_id}, ano {year}")

            # Cargar template
            template_path = os.path.join("templates", "template_format_c.xlsx")
            if not os.path.exists(template_path):
                return {
                    "status": "error",
                    "message": f"Template C no encontrado: {template_path}"
                }

            wb = load_workbook(template_path)
            ws = wb.active

            # Obtener datos del empleado para todos los meses del ano
            from database import get_payroll_by_employee_year
            records = get_payroll_by_employee_year(employee_id, year)

            if not records:
                wb.close()
                return {
                    "status": "error",
                    "message": f"No hay datos para empleado {employee_id} en {year}"
                }

            # Obtener informacion del empleado
            first_record = records[0]
            employee_name = first_record.get('name_jp', '')
            hire_date = first_record.get('hire_date', '')
            department = first_record.get('department', '')
            gender = first_record.get('gender', '')

            print(f"[INFO] Empleado: {employee_name}")
            print(f"[INFO] Registros encontrados: {len(records)}")

            # LLENAR DATOS DEL TEMPLATE

            # Row 1: Titulo
            ws['B1'] = f"賃    金    台    帳"

            # Row 6-8: Informacion del empleado
            # R6: 所属 (Department)
            ws['R6'] = department

            # AP6: 氏名 (Name)
            ws['AP6'] = employee_name

            # BJ6: 性別 (Gender)
            ws['BJ6'] = gender

            # B8: Hire date
            if hire_date:
                try:
                    hd = datetime.strptime(hire_date, '%Y-%m-%d')
                    ws['B8'] = f"{hd.year}年  {hd.month}月  {hd.day}日  雇入"
                except:
                    ws['B8'] = f"{hire_date}  雇入"

            # Organizar records por mes (1-12)
            records_by_month = {}
            for rec in records:
                period = rec.get('period', '')
                if '-' in period:
                    month = int(period.split('-')[1])
                    records_by_month[month] = rec

            # Mapeo de campos DB -> filas del template
            # Template C tiene filas diferentes
            field_mapping = {
                14: 'work_days',         # 労働日数
                16: 'work_hours',        # 労働時間
                18: 'holiday_hours',     # 休日労働時間数
                22: 'night_hours',       # 深夜残業時間数
                24: 'base_pay',          # 基本給
                26: 'overtime_pay',      # 所定時間外割増賃金
            }

            # Columnas mensuales: L(12), P(16), T(20), X(24), AB(28), AF(32), AJ(36), AN(40), AR(44), AV(48), AZ(52), BD(56)
            # Formula: 12 + (month - 1) * 4
            monthly_columns = [12 + i * 4 for i in range(12)]  # [12, 16, 20, 24, 28, 32, 36, 40, 44, 48, 52, 56]
            total_column = 60  # Column BH (合計)

            # Llenar datos por mes
            for month_idx, month in enumerate(range(1, 13)):
                col = monthly_columns[month_idx]

                if month in records_by_month:
                    rec = records_by_month[month]

                    # Llenar cada campo segun el mapeo
                    for row, field in field_mapping.items():
                        value = rec.get(field, 0)
                        if value and value != 0:
                            cell = ws.cell(row=row, column=col)
                            cell.value = value

                            # Aplicar formato segun tipo de dato
                            if field in ['work_days']:
                                cell.number_format = '0'
                            elif field in ['work_hours', 'holiday_hours', 'night_hours']:
                                cell.number_format = '0.0'
                            else:
                                cell.number_format = '#,##0'

                    # Row 28: 手当 (Allowances) - incluye commuting_allowance y otras
                    allowances = (rec.get('commuting_allowance', 0) +
                                rec.get('night_pay', 0) +
                                rec.get('holiday_pay', 0))
                    if allowances > 0:
                        ws.cell(row=28, column=col).value = allowances
                        ws.cell(row=28, column=col).number_format = '#,##0'

                    # Row 40: 小計 (Subtotal - Payments)
                    subtotal = (rec.get('base_pay', 0) +
                               rec.get('overtime_pay', 0) +
                               allowances)
                    if subtotal > 0:
                        ws.cell(row=40, column=col).value = subtotal
                        ws.cell(row=40, column=col).number_format = '#,##0'

                    # Row 46: 合計 (Total - Gross pay)
                    total_pay = subtotal  # En template C, por ahora es igual al subtotal
                    if total_pay > 0:
                        ws.cell(row=46, column=col).value = total_pay
                        ws.cell(row=46, column=col).number_format = '#,##0'

                    # Row 48: 控除額 (Deductions)
                    deductions = (rec.get('health_insurance', 0) +
                                 rec.get('care_insurance', 0) +
                                 rec.get('pension_insurance', 0) +
                                 rec.get('employment_insurance', 0) +
                                 rec.get('income_tax', 0) +
                                 rec.get('resident_tax', 0))
                    if deductions > 0:
                        ws.cell(row=48, column=col).value = deductions
                        ws.cell(row=48, column=col).number_format = '#,##0'

            # Calcular totales anuales en columna BH (60)
            for row, field in field_mapping.items():
                total = sum(records_by_month[m].get(field, 0) for m in records_by_month)
                if total > 0:
                    cell = ws.cell(row=row, column=total_column)
                    cell.value = total

                    # Aplicar formato
                    if field in ['work_days']:
                        cell.number_format = '0'
                    elif field in ['work_hours', 'holiday_hours', 'night_hours']:
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '#,##0'

            # Row 28: Total allowances
            total_allowances = sum(
                records_by_month[m].get('commuting_allowance', 0) +
                records_by_month[m].get('night_pay', 0) +
                records_by_month[m].get('holiday_pay', 0)
                for m in records_by_month
            )
            if total_allowances > 0:
                ws.cell(row=28, column=total_column).value = total_allowances
                ws.cell(row=28, column=total_column).number_format = '#,##0'

            # Row 40: Total subtotal
            total_subtotal = sum(
                records_by_month[m].get('base_pay', 0) +
                records_by_month[m].get('overtime_pay', 0) +
                records_by_month[m].get('commuting_allowance', 0) +
                records_by_month[m].get('night_pay', 0) +
                records_by_month[m].get('holiday_pay', 0)
                for m in records_by_month
            )
            if total_subtotal > 0:
                ws.cell(row=40, column=total_column).value = total_subtotal
                ws.cell(row=40, column=total_column).number_format = '#,##0'

            # Row 46: Total gross pay
            if total_subtotal > 0:
                ws.cell(row=46, column=total_column).value = total_subtotal
                ws.cell(row=46, column=total_column).number_format = '#,##0'

            # Row 48: Total deductions
            total_deductions = sum(
                records_by_month[m].get('health_insurance', 0) +
                records_by_month[m].get('care_insurance', 0) +
                records_by_month[m].get('pension_insurance', 0) +
                records_by_month[m].get('employment_insurance', 0) +
                records_by_month[m].get('income_tax', 0) +
                records_by_month[m].get('resident_tax', 0)
                for m in records_by_month
            )
            if total_deductions > 0:
                ws.cell(row=48, column=total_column).value = total_deductions
                ws.cell(row=48, column=total_column).number_format = '#,##0'

            # Guardar archivo
            if output_path is None:
                os.makedirs("outputs", exist_ok=True)
                safe_name = employee_name.replace('/', '_').replace('\\', '_')
                output_path = f"outputs/賃金台帳_{safe_name}_{year}_FormatC.xlsx"

            wb.save(output_path)
            wb.close()

            print(f"[OK] Archivo generado: {output_path}")

            return {
                "status": "success",
                "message": "Chingin Format C generado exitosamente",
                "file_path": output_path,
                "employee_id": employee_id,
                "employee_name": employee_name,
                "year": year,
                "records": len(records)
            }

        except Exception as e:
            print(f"[ERROR] generate_chingin_format_c: {e}")
            import traceback
            traceback.print_exc()
            return {
                "status": "error",
                "message": str(e)
            }

    def convert_excel_to_pdf(self, excel_path: str, pdf_path: str = None) -> dict:
        """
        Convertir archivo Excel a PDF usando win32com (MS Excel COM automation)

        Args:
            excel_path: Ruta del archivo Excel
            pdf_path: Ruta de salida PDF (opcional, se genera automaticamente)

        Returns:
            dict con status, message y pdf_path
        """
        try:
            import win32com.client
            import pythoncom

            if not os.path.exists(excel_path):
                return {
                    "status": "error",
                    "message": f"Archivo Excel no encontrado: {excel_path}"
                }

            # Generar ruta PDF si no se proporciona
            if pdf_path is None:
                pdf_path = excel_path.replace('.xlsx', '.pdf').replace('.xlsm', '.pdf')

            # Convertir rutas a absolutas
            excel_path = os.path.abspath(excel_path)
            pdf_path = os.path.abspath(pdf_path)

            print(f"[INFO] Convirtiendo Excel a PDF...")
            print(f"  Origen: {excel_path}")
            print(f"  Destino: {pdf_path}")

            # Inicializar COM
            pythoncom.CoInitialize()

            # Crear instancia de Excel
            excel = None
            wb = None
            try:
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                excel.DisplayAlerts = False

                # Abrir workbook
                wb = excel.Workbooks.Open(excel_path)

                # Exportar a PDF
                # Formato PDF = 0 (xlTypePDF)
                wb.ExportAsFixedFormat(0, pdf_path)

                print(f"[OK] PDF generado: {pdf_path}")

                return {
                    "status": "success",
                    "message": "Conversion a PDF exitosa",
                    "pdf_path": pdf_path
                }

            finally:
                # Cerrar y limpiar
                if wb:
                    wb.Close(SaveChanges=False)
                if excel:
                    excel.Quit()
                pythoncom.CoUninitialize()

        except ImportError:
            error_msg = "win32com no esta instalado. Instalar con: pip install pywin32"
            print(f"[ERROR] {error_msg}")
            return {
                "status": "error",
                "message": error_msg
            }
        except Exception as e:
            print(f"[ERROR] convert_excel_to_pdf: {e}")
            import traceback
            traceback.print_exc()
            return {
                "status": "error",
                "message": str(e)
            }


# Test
if __name__ == "__main__":
    print("="*60)
    print("🧪 TEST: Excel Processor v4 PRO")
    print("="*60)
    
    processor = ExcelProcessor()
    
    test_file = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"
    
    if os.path.exists(test_file):
        result = processor.process_file(test_file)
        print(f"\n✓ Resultado: {result}")
        
        summary = processor.get_summary()
        print(f"\n📊 RESUMEN:")
        print(f"   Total registros: {summary['total_records']}")
        print(f"   Empleados únicos: {summary['unique_employees']}")
        print(f"   Periodos: {summary['periods']}")
        print(f"   Registros en memoria: {len(processor.all_records)}")
        
        # Exportar
        print("\n📥 Exportando...")
        processor.export_to_excel_all("outputs/TEST_ALL.xlsx")
        print("   ✓ TEST_ALL.xlsx")
        
        processor.export_by_month("outputs/TEST_Por_mes.xlsx")
        print("   ✓ TEST_Por_mes.xlsx")
    else:
        print(f"❌ Archivo no encontrado: {test_file}")
