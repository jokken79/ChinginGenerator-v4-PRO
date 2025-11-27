#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de debug para ver todos los employee_ids detectados en la hoja 請負
"""
from openpyxl import load_workbook

file_path = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

wb = load_workbook(file_path, data_only=True)
ws = wb['請負']

print("=" * 80)
print("DEBUG: Employee IDs en hoja 請負")
print("=" * 80)

# Detectar bloques
employee_columns = []
max_col = ws.max_column if ws.max_column else 1200

for col in range(1, max_col + 1):
    val = ws.cell(row=2, column=col).value
    if val and '給' in str(val) and '明' in str(val) and '細' in str(val):
        employee_columns.append(col)

print(f"\n[INFO] Encontrados {len(employee_columns)} bloques")
print(f"[INFO] Columnas: {employee_columns[:20]}...")

# Analizar cada bloque
ids_03 = []
ids_other = []
ids_empty = []
ids_invalid = []

for idx, start_col in enumerate(employee_columns):
    # Buscar employee_id en diferentes posiciones
    positions = [
        (6, 8),   # Fila 6, Col+8 (posición actual)
        (6, 9),   # Fila 6, Col+9
        (6, 7),   # Fila 6, Col+7
        (5, 8),   # Fila 5, Col+8
        (7, 8),   # Fila 7, Col+8
    ]

    found_id = None
    found_pos = None

    for row_offset, col_offset in positions:
        emp_id = ws.cell(row=row_offset, column=start_col + col_offset).value
        if emp_id:
            emp_id_str = str(emp_id).strip()
            if emp_id_str.isdigit() and len(emp_id_str) >= 6:
                found_id = emp_id_str
                found_pos = (row_offset, col_offset)
                break

    # Obtener nombre para referencia
    name = ws.cell(row=8, column=start_col + 1).value

    if found_id:
        if found_id.startswith('03'):
            ids_03.append((found_id, name, start_col, found_pos))
        else:
            ids_other.append((found_id, name, start_col, found_pos))
    elif idx < 10:  # Solo mostrar los primeros 10 vacíos
        ids_empty.append((start_col, name))

print(f"\n" + "=" * 80)
print("RESUMEN DE IDs:")
print("=" * 80)
print(f"  IDs 03xxxx: {len(ids_03)}")
print(f"  Otros IDs: {len(ids_other)}")
print(f"  Sin ID válido: {len(employee_columns) - len(ids_03) - len(ids_other)}")

if ids_03:
    print(f"\n[INFO] Primeros 10 IDs 03xxxx:")
    for emp_id, name, col, pos in ids_03[:10]:
        print(f"  - Col {col}: ID={emp_id}, Nombre={name}, Pos={pos}")

if ids_other:
    print(f"\n[INFO] Primeros 10 otros IDs (NO 03xxxx):")
    for emp_id, name, col, pos in ids_other[:10]:
        print(f"  - Col {col}: ID={emp_id}, Nombre={name}, Pos={pos}")

if ids_empty:
    print(f"\n[INFO] Primeros bloques sin ID válido:")
    for col, name in ids_empty[:10]:
        print(f"  - Col {col}: Nombre={name}")
        # Mostrar valores alrededor de la posición esperada
        print(f"    Fila 6: ", end="")
        for c_off in range(7, 11):
            val = ws.cell(row=6, column=col + c_off).value
            print(f"Col+{c_off}={val} | ", end="")
        print()

wb.close()

print(f"\n" + "=" * 80)
print("CONCLUSION:")
print("=" * 80)
if len(ids_03) >= 50:
    print(f"[OK] Se encontraron {len(ids_03)} empleados 03xxxx")
else:
    print(f"[ATENCION] Solo se encontraron {len(ids_03)} empleados 03xxxx")
    print(f"  El usuario reporta ~50 empleados")
    print(f"  Diferencia: {50 - len(ids_03)} empleados faltantes")
