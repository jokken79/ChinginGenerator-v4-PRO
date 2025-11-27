#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook

file_path = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

wb = load_workbook(file_path, data_only=True)
ws = wb['totalChin']

print("Buscando empleados 請負社員 (IDs 03xxxx) en hoja totalChin:")
print("=" * 80)

# Headers
headers = []
for col in range(1, 10):
    headers.append(ws.cell(row=1, column=col).value)
print(f"Headers: {headers[:6]}")
print()

# Buscar empleados con ID que empiece con 03
ukeoi_count = 0
haken_count = 0
other_count = 0

for row in range(2, min(ws.max_row + 1, 500)):
    emp_id = ws.cell(row=row, column=2).value
    if emp_id:
        emp_id_str = str(emp_id).strip()
        name = ws.cell(row=row, column=4).value
        dispatch = ws.cell(row=row, column=6).value

        if emp_id_str.startswith('03'):
            ukeoi_count += 1
            if ukeoi_count <= 5:
                print(f"  ✅ 請負社員 - ID: {emp_id_str}, Nombre: {name}, Lugar: {dispatch}")
        elif emp_id_str.startswith(('20', '21', '22', '23', '24', '25')):
            haken_count += 1
        else:
            other_count += 1

print()
print("=" * 80)
print(f"Resumen en totalChin:")
print(f"  請負社員 (03xxxx): {ukeoi_count}")
print(f"  派遣社員 (20-25xxxx): {haken_count}")
print(f"  Otros: {other_count}")
print(f"  Total procesado: {ukeoi_count + haken_count + other_count}")

wb.close()
