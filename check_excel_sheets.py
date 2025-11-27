#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import os

file_path = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

if os.path.exists(file_path):
    print(f"Archivo encontrado: {os.path.basename(file_path)}")
    print("-" * 60)

    wb = load_workbook(file_path, data_only=True)
    print(f"\nHojas disponibles en el archivo:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        print(f"  {idx}. {sheet_name} ({ws.max_row} filas, {ws.max_column} columnas)")

    # Buscar hoja 請負
    ukeoi_sheet = None
    for name in wb.sheetnames:
        if '請負' in name:
            ukeoi_sheet = name
            break

    if ukeoi_sheet:
        print(f"\n✅ Hoja 請負 encontrada: '{ukeoi_sheet}'")
        ws = wb[ukeoi_sheet]

        # Leer headers
        print(f"\nPrimeros 10 headers de la hoja '{ukeoi_sheet}':")
        for col in range(1, 11):
            header = ws.cell(row=1, column=col).value
            print(f"  Col {col}: {header}")

        # Contar filas con datos
        count = 0
        for row in range(2, min(ws.max_row + 1, 20)):
            emp_id = ws.cell(row=row, column=2).value
            if emp_id and str(emp_id).strip():
                count += 1
                if count <= 3:
                    name = ws.cell(row=row, column=4).value
                    print(f"  Fila {row}: ID={emp_id}, Nombre={name}")

        print(f"\nTotal filas con datos (primeras 18): {count}")
    else:
        print("\n❌ No se encontró hoja con '請負' en el nombre")

    wb.close()
else:
    print(f"❌ Archivo no encontrado: {file_path}")
