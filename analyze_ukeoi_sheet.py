#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook

file_path = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

wb = load_workbook(file_path, data_only=True)

# Comparar estructura de totalChin vs 請負
print("=" * 80)
print("COMPARACIÓN: totalChin vs 請負")
print("=" * 80)

# Hoja totalChin
ws_total = wb['totalChin']
print("\n[totalChin] - Primera fila (headers):")
for col in range(1, 11):
    val = ws_total.cell(row=1, column=col).value
    print(f"  Col {col}: {val}")

print(f"\n[totalChin] - Fila 2 (primer registro de datos):")
for col in range(1, 8):
    val = ws_total.cell(row=2, column=col).value
    print(f"  Col {col}: {val}")

# Hoja 請負
ws_ukeoi = wb['請負']
print("\n" + "=" * 80)
print("[請負] - Estructura de las primeras 30 filas:")
print("=" * 80)

for row in range(1, 31):
    # Mostrar primeras 5 columnas de cada fila
    values = []
    for col in range(1, 6):
        val = ws_ukeoi.cell(row=row, column=col).value
        if val:
            val_str = str(val)[:30]  # Limitar longitud
            values.append(f"Col{col}={val_str}")

    if values:
        print(f"  Fila {row:2d}: {' | '.join(values)}")

# Buscar fila con headers tipo "従業員番号"
print("\n" + "=" * 80)
print("Buscando fila de headers en 請負:")
print("=" * 80)

header_row = None
for row in range(1, 50):
    val = ws_ukeoi.cell(row=row, column=2).value
    if val and '従業員番号' in str(val):
        header_row = row
        break

if header_row:
    print(f"\n✅ Headers encontrados en fila {header_row}:")
    for col in range(1, 11):
        val = ws_ukeoi.cell(row=header_row, column=col).value
        print(f"  Col {col}: {val}")

    print(f"\n[請負] - Fila {header_row + 1} (primer registro de datos):")
    for col in range(1, 8):
        val = ws_ukeoi.cell(row=header_row + 1, column=col).value
        print(f"  Col {col}: {val}")
else:
    print("\n❌ No se encontró fila con headers '従業員番号'")

wb.close()
