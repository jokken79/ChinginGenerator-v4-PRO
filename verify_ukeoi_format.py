#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
import io
from openpyxl import load_workbook

# Fix encoding for Windows console
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

wb = load_workbook(file_path, data_only=True)

print("=" * 100)
print("VERIFICACIÓN: Comparando formato totalChin vs 請負")
print("=" * 100)

# TOTALCHIN
ws_total = wb['totalChin']
print(f"\n[totalChin] Dimensiones: {ws_total.max_row} filas x {ws_total.max_column} columnas")
print(f"\n[totalChin] Headers (fila 1, primeras 10 columnas):")
for col in range(1, 11):
    h = ws_total.cell(row=1, column=col).value
    print(f"  Col {col:2d}: {h}")

print(f"\n[totalChin] Primer registro (fila 2, primeras 10 columnas):")
for col in range(1, 11):
    v = ws_total.cell(row=2, column=col).value
    print(f"  Col {col:2d}: {v}")

# UKEOI
print("\n" + "=" * 100)
ws_ukeoi = wb['請負']
print(f"\n[請負] Dimensiones: {ws_ukeoi.max_row} filas x {ws_ukeoi.max_column} columnas")
print(f"\n[請負] Fila 1 (primeras 10 columnas):")
for col in range(1, 11):
    h = ws_ukeoi.cell(row=1, column=col).value
    print(f"  Col {col:2d}: {h}")

print(f"\n[請負] Fila 2 (primeras 10 columnas):")
for col in range(1, 11):
    v = ws_ukeoi.cell(row=2, column=col).value
    print(f"  Col {col:2d}: {v}")

# BUSCAR HEADERS EN 請負
print("\n" + "=" * 100)
print("BUSCANDO fila de headers en 請負 (buscando '従業員番号'):")
print("=" * 100)

header_row_found = None
for row in range(1, 50):
    for col in range(1, 20):
        val = ws_ukeoi.cell(row=row, column=col).value
        if val and '従業員番号' in str(val):
            header_row_found = row
            print(f"\n✅ ENCONTRADO '従業員番号' en: Fila {row}, Col {col}")
            break
    if header_row_found:
        break

if header_row_found:
    print(f"\n[請負] Headers encontrados en FILA {header_row_found} (primeras 10 columnas):")
    for col in range(1, 11):
        h = ws_ukeoi.cell(row=header_row_found, column=col).value
        print(f"  Col {col:2d}: {h}")

    print(f"\n[請負] Primer registro de datos en FILA {header_row_found + 1} (primeras 10 columnas):")
    for col in range(1, 11):
        v = ws_ukeoi.cell(row=header_row_found + 1, column=col).value
        print(f"  Col {col:2d}: {v}")

    # Contar empleados con ID 03xxxx
    print(f"\n[請負] Buscando empleados con ID 03xxxx (primeras 50 filas):")
    count = 0
    for row in range(header_row_found + 1, min(ws_ukeoi.max_row + 1, header_row_found + 51)):
        emp_id = ws_ukeoi.cell(row=row, column=2).value  # Col 2 = 従業員番号
        if emp_id and str(emp_id).strip().startswith('03'):
            name = ws_ukeoi.cell(row=row, column=4).value  # Col 4 = 氏名
            count += 1
            if count <= 5:
                print(f"  ✅ Fila {row}: ID={emp_id}, Nombre={name}")

    print(f"\n  Total empleados 03xxxx encontrados: {count}")

    print("\n" + "=" * 100)
    print("CONCLUSIÓN:")
    print("=" * 100)
    if header_row_found == 1:
        print("✅ La hoja 請負 tiene EXACTAMENTE el mismo formato que totalChin")
        print("   (Headers en fila 1, datos desde fila 2)")
    else:
        print(f"⚠️  La hoja 請負 tiene headers en FILA {header_row_found}, no en fila 1")
        print(f"   Necesita ajuste en el código para saltar las primeras {header_row_found - 1} filas")
else:
    print("\n❌ NO se encontró la columna '従業員番号' en las primeras 50 filas")
    print("   La hoja 請負 tiene un formato DIFERENTE a totalChin")

wb.close()
