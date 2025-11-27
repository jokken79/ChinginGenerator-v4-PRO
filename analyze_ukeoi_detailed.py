#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import re

file_path = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

wb = load_workbook(file_path, data_only=True)
ws = wb['請負']

print("=" * 100)
print("ANÁLISIS DETALLADO: Estructura de la hoja 請負")
print("=" * 100)

# Detectar bloques de empleados buscando el encabezado "給　料　支　払　明　細　書"
employee_blocks = []

for col in range(1, min(ws.max_column + 1, 100), 5):  # Revisar cada 5 columnas, primeras 100
    for check_col in range(col, min(col + 5, ws.max_column + 1)):
        val = ws.cell(row=2, column=check_col).value
        if val and '給' in str(val) and '明' in str(val) and '細' in str(val):
            employee_blocks.append(check_col)
            break

print(f"\n✅ Encontrados {len(employee_blocks)} bloques de empleados")
print(f"Columnas de inicio: {employee_blocks[:10]}")  # Mostrar primeros 10

# Analizar los primeros 3 bloques en detalle
for idx, start_col in enumerate(employee_blocks[:3]):
    print(f"\n" + "=" * 100)
    print(f"BLOQUE #{idx + 1} - Columna de inicio: {start_col}")
    print("=" * 100)

    # Mostrar todas las filas con datos en un rango de 15 columnas
    print(f"\nDatos del bloque (filas 1-60, columnas {start_col} a {start_col+14}):")

    for row in range(1, 61):
        row_data = {}
        for offset in range(15):
            col = start_col + offset
            val = ws.cell(row=row, column=col).value
            if val:
                row_data[offset] = str(val).strip()[:40]  # Limitar longitud

        if row_data:
            data_str = " | ".join([f"Col+{k}={v}" for k, v in row_data.items()])
            print(f"  Fila {row:2d}: {data_str}")

    # Intentar extraer información estructurada
    print(f"\n" + "-" * 100)
    print("INFORMACIÓN ESTRUCTURADA EXTRAÍDA:")
    print("-" * 100)

    # Buscar período
    for r in range(1, 10):
        for c_offset in range(10):
            val = ws.cell(row=r, column=start_col + c_offset).value
            if val and re.search(r'\d{4}年\d+月分', str(val)):
                print(f"  Período: Fila {r}, Col+{c_offset} = {val}")
                break

    # Buscar nombre
    for r in range(1, 15):
        for c_offset in range(10):
            val = ws.cell(row=r, column=start_col + c_offset).value
            if val and '氏名' in str(val):
                print(f"  Nombre: Fila {r}, Col+{c_offset} = {val}")
                break

    # Buscar días trabajados
    for r in range(9, 13):
        for c_offset in range(10):
            label = ws.cell(row=r, column=start_col + c_offset).value
            if label and '日' in str(label) and '数' in str(label):
                value = ws.cell(row=r, column=start_col + c_offset + 1).value
                print(f"  Días: Fila {r}, Col+{c_offset} (label), Col+{c_offset+1} (value={value})")
                break

    # Buscar horas
    for r in range(12, 17):
        for c_offset in range(10):
            label = ws.cell(row=r, column=start_col + c_offset).value
            if label and '時' in str(label) and '間' in str(label):
                value = ws.cell(row=r, column=start_col + c_offset + 1).value
                print(f"  Horas ({label[:15]}): Fila {r}, Col+{c_offset} (label), Col+{c_offset+1} (value={value})")

    # Buscar salarios
    for r in range(15, 35):
        for c_offset in range(10):
            label = ws.cell(row=r, column=start_col + c_offset).value
            value = ws.cell(row=r, column=start_col + c_offset + 1).value
            if label and value and isinstance(value, (int, float)) and value > 0:
                if '給' in str(label) or '手' in str(label):
                    print(f"  Pago ({str(label)[:15]}): Fila {r}, Col+{c_offset} (label), Col+{c_offset+1} (value={value})")

    # Buscar total
    for r in range(25, 35):
        for c_offset in range(10):
            label = ws.cell(row=r, column=start_col + c_offset).value
            if label and '合' in str(label) and '計' in str(label):
                value = ws.cell(row=r, column=start_col + c_offset + 1).value
                print(f"  Total: Fila {r}, Col+{c_offset} (label), Col+{c_offset+1} (value={value})")
                break

wb.close()
