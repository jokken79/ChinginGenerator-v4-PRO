#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import re

file_path = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

wb = load_workbook(file_path, data_only=True)
ws = wb['請負']

print("Analizando estructura de la hoja 請負...")
print("=" * 80)

# La hoja 請負 tiene un formato donde cada empleado ocupa un bloque vertical
# Necesito encontrar el patrón que separa a cada empleado

# Buscar bloques que empiecen con "給　料　支　払　明　細　書"
employee_blocks = []
current_block_start = None

for row in range(1, min(ws.max_row + 1, 100)):
    val_b = ws.cell(row=row, column=2).value
    if val_b and '給　料　支　払　明　細　書' in str(val_b):
        if current_block_start:
            employee_blocks.append((current_block_start, row - 1))
        current_block_start = row

if current_block_start:
    employee_blocks.append((current_block_start, min(ws.max_row, current_block_start + 50)))

print(f"\nEncontrados {len(employee_blocks)} bloques de empleados")
print(f"Bloques: {employee_blocks[:3]}")

# Analizar el primer bloque
if employee_blocks:
    start, end = employee_blocks[0]
    print(f"\n" + "=" * 80)
    print(f"ANALIZANDO PRIMER BLOQUE (filas {start} a {end}):")
    print("=" * 80)

    emp_data = {}

    for row in range(start, min(end + 1, start + 40)):
        # Leer valores de las columnas relevantes
        vals = {}
        for col in range(1, 10):
            v = ws.cell(row=row, column=col).value
            if v:
                vals[col] = str(v).strip()

        # Mostrar fila si tiene datos
        if vals:
            print(f"  Fila {row}: {vals}")

            # Extraer información clave
            for col, text in vals.items():
                # Período
                if re.search(r'\d{4}年\d+月分', text):
                    emp_data['period'] = text
                    print(f"    -> PERIODO: {text}")

                # Nombre en katakana (fila 7, col 3)
                if row == start + 6 and col == 3:
                    emp_data['name_kana'] = text
                    print(f"    -> NOMBRE KANA: {text}")

                # Nombre kanji (fila 8, col 3)
                if row == start + 7 and col == 3 and '氏名' in text:
                    # Extraer nombre después de "氏名"
                    match = re.search(r'氏名\s*(.+)', text)
                    if match:
                        emp_data['name'] = match.group(1).strip()
                        print(f"    -> NOMBRE: {emp_data['name']}")

                # Días trabajados
                if '労' in text and '働' in text and '日' in text and '数' in text:
                    next_val = ws.cell(row=row, column=col+1).value
                    if next_val:
                        emp_data['work_days'] = next_val
                        print(f"    -> DIAS TRABAJADOS: {next_val}")

                # Horas trabajadas
                if '労' in text and '働' in text and '時' in text and '間' in text:
                    next_val = ws.cell(row=row, column=col+1).value
                    if next_val:
                        emp_data['work_hours'] = next_val
                        print(f"    -> HORAS TRABAJADAS: {next_val}")

                # Horas extra
                if '所 定 時 間 外' in text or ('時 間 外' in text and '労 働' in text):
                    next_val = ws.cell(row=row, column=col+1).value
                    if next_val:
                        emp_data['overtime_hours'] = next_val
                        print(f"    -> HORAS EXTRA: {next_val}")

                # Horas nocturnas
                if '深 夜' in text and '労 働' in text:
                    next_val = ws.cell(row=row, column=col+1).value
                    if next_val:
                        emp_data['night_hours'] = next_val
                        print(f"    -> HORAS NOCTURNAS: {next_val}")

                # Salario base
                if '基' in text and '本' in text and '給' in text:
                    next_val = ws.cell(row=row, column=col+1).value
                    if next_val and isinstance(next_val, (int, float)):
                        emp_data['base_pay'] = next_val
                        print(f"    -> SALARIO BASE: {next_val}")

                # Total
                if '合' in text and '計' in text and col == 3:
                    next_val = ws.cell(row=row, column=col+1).value
                    if next_val and isinstance(next_val, (int, float)):
                        emp_data['total_pay'] = next_val
                        print(f"    -> TOTAL: {next_val}")

    print(f"\n" + "=" * 80)
    print(f"DATOS EXTRAIDOS DEL PRIMER EMPLEADO:")
    print("=" * 80)
    for key, val in emp_data.items():
        print(f"  {key}: {val}")

wb.close()
