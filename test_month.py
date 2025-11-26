#!/usr/bin/env python3
"""Test rápido para verificar agrupación por mes"""
from excel_processor import ExcelProcessor
import os

processor = ExcelProcessor()
test_file = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

print("Procesando archivo...")
if os.path.exists(test_file):
    result = processor.process_file(test_file)
    print(f"Resultado: {result}")
    print(f"Registros en memoria: {len(processor.all_records)}")
    
    # Exportar
    print("\nExportando...")
    processor.export_by_month("outputs/TEST_Por_mes.xlsx")
    print("Exportado: outputs/TEST_Por_mes.xlsx")
    
    # Verificar hojas
    from openpyxl import load_workbook
    wb = load_workbook("outputs/TEST_Por_mes.xlsx")
    print(f"\n✓ Hojas generadas: {wb.sheetnames}")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"  {sheet}: {ws.max_row - 1} empleados")
    wb.close()
else:
    print("Archivo no encontrado")
