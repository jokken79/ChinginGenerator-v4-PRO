#!/usr/bin/env python3
"""
賃金台帳 Generator v4 PRO - Database Module
Base de datos SQLite para persistencia, auditoría y backups
"""

import sqlite3
import os
import json
import hashlib
from datetime import datetime, timedelta
from contextlib import contextmanager
from typing import Optional, List, Dict, Any

# Ruta de la base de datos
# En Docker usa /app/data, localmente usa el directorio actual
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data") if os.path.exists(os.path.join(BASE_DIR, "data")) else BASE_DIR
os.makedirs(DATA_DIR, exist_ok=True)

DB_PATH = os.path.join(DATA_DIR, "chingin_data.db")
BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backups")


def get_db_path():
    return DB_PATH


@contextmanager
def get_connection():
    """Context manager para conexiones a la base de datos"""
    conn = sqlite3.connect(DB_PATH, timeout=30, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    # Habilitar WAL mode para mejor concurrencia
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    try:
        yield conn
        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()


def init_database():
    """Inicializa la base de datos con todas las tablas"""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # ========================================
        # TABLA: employees (従業員)
        # ========================================
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id TEXT UNIQUE NOT NULL,
                name_roman TEXT,
                name_jp TEXT,
                hourly_rate REAL,
                department TEXT,
                position TEXT,
                hire_date TEXT,
                status TEXT DEFAULT 'active',
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # ========================================
        # TABLA: payroll_records (賃金記録)
        # ========================================
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS payroll_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id TEXT NOT NULL,
                period TEXT NOT NULL,
                period_start TEXT,
                period_end TEXT,
                work_days INTEGER DEFAULT 0,
                work_hours REAL DEFAULT 0,
                overtime_hours REAL DEFAULT 0,
                night_hours REAL DEFAULT 0,
                holiday_hours REAL DEFAULT 0,
                base_pay REAL DEFAULT 0,
                overtime_pay REAL DEFAULT 0,
                night_pay REAL DEFAULT 0,
                holiday_pay REAL DEFAULT 0,
                commuting_allowance REAL DEFAULT 0,
                total_pay REAL DEFAULT 0,
                health_insurance REAL DEFAULT 0,
                pension REAL DEFAULT 0,
                employment_insurance REAL DEFAULT 0,
                income_tax REAL DEFAULT 0,
                resident_tax REAL DEFAULT 0,
                deduction_total REAL DEFAULT 0,
                net_pay REAL DEFAULT 0,
                source_file TEXT,
                raw_data TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(employee_id, period),
                FOREIGN KEY (employee_id) REFERENCES employees(employee_id)
            )
        """)
        
        # ========================================
        # TABLA: audit_log (監査ログ)
        # ========================================
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                table_name TEXT,
                record_id TEXT,
                old_value TEXT,
                new_value TEXT,
                user_info TEXT,
                ip_address TEXT,
                details TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # ========================================
        # TABLA: backups (バックアップ)
        # ========================================
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS backups (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                filepath TEXT NOT NULL,
                file_hash TEXT NOT NULL,
                file_size INTEGER,
                backup_type TEXT DEFAULT 'auto',
                description TEXT,
                is_valid INTEGER DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # ========================================
        # TABLA: processed_files (処理済みファイル)
        # ========================================
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS processed_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                filename TEXT NOT NULL,
                filepath TEXT,
                file_hash TEXT,
                file_size INTEGER,
                records_count INTEGER DEFAULT 0,
                status TEXT DEFAULT 'success',
                error_message TEXT,
                processed_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # ========================================
        # TABLA: settings (設定)
        # ========================================
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT,
                description TEXT,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Insertar configuraciones por defecto
        default_settings = [
            ('auto_backup_enabled', 'true', 'Habilitar backup automático'),
            ('auto_backup_interval_hours', '24', 'Intervalo de backup en horas'),
            ('max_backups_keep', '30', 'Número máximo de backups a mantener'),
            ('integrity_check_enabled', 'true', 'Verificar integridad SHA256'),
            ('audit_log_enabled', 'true', 'Habilitar log de auditoría'),
        ]
        
        for key, value, desc in default_settings:
            cursor.execute("""
                INSERT OR IGNORE INTO settings (key, value, description)
                VALUES (?, ?, ?)
            """, (key, value, desc))
        
        # ========================================
        # TABLA: haken_employees (派遣社員マスター)
        # ========================================
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS haken_employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id TEXT UNIQUE NOT NULL,
                status TEXT,
                dispatch_id TEXT,
                dispatch_company TEXT,
                department TEXT,
                line TEXT,
                job_description TEXT,
                name TEXT,
                name_kana TEXT,
                gender TEXT,
                nationality TEXT,
                birth_date TEXT,
                age INTEGER,
                hourly_rate REAL,
                hourly_rate_history TEXT,
                billing_rate REAL,
                billing_history TEXT,
                profit_margin REAL,
                standard_salary REAL,
                health_insurance REAL,
                care_insurance REAL,
                pension REAL,
                visa_expiry TEXT,
                visa_alert TEXT,
                visa_type TEXT,
                postal_code TEXT,
                address TEXT,
                apartment TEXT,
                move_in_date TEXT,
                hire_date TEXT,
                synced_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # ========================================
        # TABLA: ukeoi_employees (請負社員マスター)
        # ========================================
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS ukeoi_employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id TEXT UNIQUE NOT NULL,
                status TEXT,
                job_type TEXT,
                name TEXT,
                name_kana TEXT,
                gender TEXT,
                nationality TEXT,
                birth_date TEXT,
                age INTEGER,
                hourly_rate REAL,
                hourly_rate_history TEXT,
                standard_salary REAL,
                health_insurance REAL,
                care_insurance REAL,
                pension REAL,
                commute_distance REAL,
                transport_fee REAL,
                profit_margin REAL,
                visa_expiry TEXT,
                visa_alert TEXT,
                visa_type TEXT,
                postal_code TEXT,
                address TEXT,
                apartment TEXT,
                move_in_date TEXT,
                hire_date TEXT,
                resignation_date TEXT,
                move_out_date TEXT,
                social_insurance TEXT,
                account_name TEXT,
                synced_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        # Crear indices para mejor rendimiento
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_payroll_employee ON payroll_records(employee_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_payroll_period ON payroll_records(period)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_audit_action ON audit_log(action)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_audit_date ON audit_log(created_at)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_haken_employee_id ON haken_employees(employee_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_ukeoi_employee_id ON ukeoi_employees(employee_id)")

        # Migración: Agregar columna commuting_allowance si no existe
        try:
            cursor.execute("SELECT commuting_allowance FROM payroll_records LIMIT 1")
        except sqlite3.OperationalError:
            print("[INFO] Agregando columna commuting_allowance a payroll_records...")
            cursor.execute("ALTER TABLE payroll_records ADD COLUMN commuting_allowance REAL DEFAULT 0")
            print("[OK] Columna commuting_allowance agregada")

        conn.commit()
        print("[OK] Base de datos inicializada correctamente")


# ========================================
# FUNCIONES DE EMPLEADOS
# ========================================

def upsert_employee(employee_data: Dict) -> int:
    """Insertar o actualizar empleado"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO employees (employee_id, name_roman, name_jp, hourly_rate)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(employee_id) DO UPDATE SET
                name_roman = excluded.name_roman,
                name_jp = excluded.name_jp,
                hourly_rate = COALESCE(excluded.hourly_rate, hourly_rate),
                updated_at = CURRENT_TIMESTAMP
        """, (
            employee_data.get('employee_id'),
            employee_data.get('name_roman'),
            employee_data.get('name_jp'),
            employee_data.get('hourly_rate')
        ))
        
        return cursor.lastrowid


def get_all_employees() -> List[Dict]:
    """Obtener todos los empleados"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM employees WHERE status = 'active' ORDER BY employee_id")
        return [dict(row) for row in cursor.fetchall()]


def get_employee(employee_id: str) -> Optional[Dict]:
    """Obtener empleado por ID"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM employees WHERE employee_id = ?", (employee_id,))
        row = cursor.fetchone()
        return dict(row) if row else None


# ========================================
# FUNCIONES DE NÓMINA
# ========================================

def save_payroll_record(record: Dict) -> int:
    """Guardar registro de nómina"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Primero asegurar que el empleado existe (inline para evitar bloqueo)
        cursor.execute("""
            INSERT INTO employees (employee_id, name_roman, name_jp, hourly_rate)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(employee_id) DO UPDATE SET
                name_roman = COALESCE(excluded.name_roman, employees.name_roman),
                name_jp = COALESCE(excluded.name_jp, employees.name_jp),
                hourly_rate = COALESCE(excluded.hourly_rate, employees.hourly_rate),
                updated_at = CURRENT_TIMESTAMP
        """, (
            record.get('employee_id'),
            record.get('name_roman'),
            record.get('name_jp'),
            record.get('hourly_rate')
        ))
        
        cursor.execute("""
            INSERT INTO payroll_records (
                employee_id, period, period_start, period_end,
                work_days, work_hours, overtime_hours, night_hours, holiday_hours,
                base_pay, overtime_pay, night_pay, holiday_pay, commuting_allowance, total_pay,
                health_insurance, pension, employment_insurance,
                income_tax, resident_tax, deduction_total, net_pay,
                source_file, raw_data
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(employee_id, period) DO UPDATE SET
                work_days = excluded.work_days,
                work_hours = excluded.work_hours,
                overtime_hours = excluded.overtime_hours,
                night_hours = excluded.night_hours,
                holiday_hours = excluded.holiday_hours,
                base_pay = excluded.base_pay,
                overtime_pay = excluded.overtime_pay,
                night_pay = excluded.night_pay,
                holiday_pay = excluded.holiday_pay,
                commuting_allowance = excluded.commuting_allowance,
                total_pay = excluded.total_pay,
                health_insurance = excluded.health_insurance,
                pension = excluded.pension,
                employment_insurance = excluded.employment_insurance,
                income_tax = excluded.income_tax,
                resident_tax = excluded.resident_tax,
                deduction_total = excluded.deduction_total,
                net_pay = excluded.net_pay,
                source_file = excluded.source_file,
                raw_data = excluded.raw_data,
                updated_at = CURRENT_TIMESTAMP
        """, (
            record.get('employee_id'),
            record.get('period'),
            record.get('period_start'),
            record.get('period_end'),
            record.get('work_days', 0),
            record.get('work_hours', 0),
            record.get('overtime_hours', 0),
            record.get('night_hours', 0),
            record.get('holiday_hours', 0),
            record.get('base_pay', 0),
            record.get('overtime_pay', 0),
            record.get('night_pay', 0),
            record.get('holiday_pay', 0),
            record.get('commuting_allowance', 0),
            record.get('total_pay', 0),
            record.get('health_insurance', 0),
            record.get('pension', 0),
            record.get('employment_insurance', 0),
            record.get('income_tax', 0),
            record.get('resident_tax', 0),
            record.get('deduction_total', 0),
            record.get('net_pay', 0),
            record.get('source_file'),
            json.dumps(record, ensure_ascii=False, default=str)
        ))
        
        # Log de auditoría (inline para evitar bloqueo)
        cursor.execute("""
            INSERT INTO audit_log (action, table_name, record_id, old_value, new_value, details)
            VALUES (?, ?, ?, ?, ?, ?)
        """, ('INSERT_PAYROLL', 'payroll_records', record.get('employee_id'), 
              None, json.dumps(record, ensure_ascii=False, default=str), None))
        
        return cursor.lastrowid


def get_payroll_by_employee(employee_id: str) -> List[Dict]:
    """Obtener nóminas de un empleado"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT * FROM payroll_records 
            WHERE employee_id = ? 
            ORDER BY period DESC
        """, (employee_id,))
        return [dict(row) for row in cursor.fetchall()]


def get_payroll_by_employee_year(employee_id: str, year: int) -> List[Dict]:
    """Obtener nominas de un empleado para un ano especifico"""
    import re
    with get_connection() as conn:
        cursor = conn.cursor()
        # Buscar por formato japonés "2025年..." o formato "2025-%"
        cursor.execute("""
            SELECT pr.*, e.name_roman, e.name_jp, e.hire_date, e.department
            FROM payroll_records pr
            LEFT JOIN employees e ON pr.employee_id = e.employee_id
            WHERE pr.employee_id = ?
              AND (pr.period LIKE ? OR pr.period LIKE ?)
            ORDER BY pr.period ASC
        """, (employee_id, f"{year}年%", f"{year}-%"))

        records = [dict(row) for row in cursor.fetchall()]

        # Normalizar el periodo a formato "YYYY-MM" para procesamiento
        for record in records:
            period = record.get('period', '')
            # Si ya está en formato YYYY-MM, dejarlo
            if re.match(r'^\d{4}-\d{2}$', period):
                continue
            # Extraer año y mes del formato japonés "2025年2月分..."
            match = re.search(r'(\d{4})年(\d{1,2})月', period)
            if match:
                year_val = match.group(1)
                month_val = match.group(2).zfill(2)
                record['period'] = f"{year_val}-{month_val}"

        # Obtener datos adicionales de tablas maestras (gender, birth_date)
        if records:
            emp_id = records[0].get('employee_id')
            # Buscar en haken_employees
            cursor.execute("""
                SELECT gender, birth_date FROM haken_employees WHERE employee_id = ?
            """, (emp_id,))
            haken_row = cursor.fetchone()
            if haken_row:
                haken_data = dict(haken_row)
                for record in records:
                    record['gender'] = haken_data.get('gender', '')
                    record['birth_date'] = haken_data.get('birth_date', '')
            else:
                # Buscar en ukeoi_employees
                cursor.execute("""
                    SELECT gender, birth_date FROM ukeoi_employees WHERE employee_id = ?
                """, (emp_id,))
                ukeoi_row = cursor.fetchone()
                if ukeoi_row:
                    ukeoi_data = dict(ukeoi_row)
                    for record in records:
                        record['gender'] = ukeoi_data.get('gender', '')
                        record['birth_date'] = ukeoi_data.get('birth_date', '')

        return records


def get_payroll_by_period(period: str) -> List[Dict]:
    """Obtener nóminas de un periodo"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT pr.*, e.name_roman, e.name_jp
            FROM payroll_records pr
            LEFT JOIN employees e ON pr.employee_id = e.employee_id
            WHERE pr.period = ?
            ORDER BY pr.employee_id
        """, (period,))
        return [dict(row) for row in cursor.fetchall()]


def get_all_payroll_records() -> List[Dict]:
    """Obtener todos los registros de nómina"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT pr.*, e.name_roman, e.name_jp
            FROM payroll_records pr
            LEFT JOIN employees e ON pr.employee_id = e.employee_id
            ORDER BY pr.period DESC, pr.employee_id
        """)
        return [dict(row) for row in cursor.fetchall()]


def get_periods() -> List[str]:
    """Obtener lista de periodos únicos"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT DISTINCT period FROM payroll_records ORDER BY period DESC")
        return [row[0] for row in cursor.fetchall()]


# ========================================
# FUNCIONES DE AUDITORÍA
# ========================================

def log_audit(action: str, table_name: str = None, record_id: str = None,
              old_value: str = None, new_value: str = None, details: str = None):
    """Registrar acción en log de auditoría"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO audit_log (action, table_name, record_id, old_value, new_value, details)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (action, table_name, record_id, old_value, new_value, details))


def get_audit_log(limit: int = 100, action_filter: str = None) -> List[Dict]:
    """Obtener log de auditoría"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        if action_filter:
            cursor.execute("""
                SELECT * FROM audit_log 
                WHERE action LIKE ? 
                ORDER BY created_at DESC LIMIT ?
            """, (f"%{action_filter}%", limit))
        else:
            cursor.execute("""
                SELECT * FROM audit_log 
                ORDER BY created_at DESC LIMIT ?
            """, (limit,))
        
        return [dict(row) for row in cursor.fetchall()]


def clear_all_data():
    """Borrar TODOS los datos de la base de datos (excepto backups y configuración)"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Contar registros antes de borrar
        cursor.execute("SELECT COUNT(*) FROM payroll_records")
        payroll_count = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM employees")
        employee_count = cursor.fetchone()[0]
        
        # Borrar datos
        cursor.execute("DELETE FROM payroll_records")
        cursor.execute("DELETE FROM employees")
        cursor.execute("DELETE FROM processed_files")
        
        # Registrar en auditoría
        cursor.execute("""
            INSERT INTO audit_log (action, table_name, details)
            VALUES (?, ?, ?)
        """, ('CLEAR_ALL_DATA', 'all', f'Borrados {payroll_count} registros de nómina y {employee_count} empleados'))
        
        return {
            "payroll_deleted": payroll_count,
            "employees_deleted": employee_count,
            "status": "success"
        }


# ========================================
# FUNCIONES DE BACKUP E INTEGRIDAD
# ========================================

def calculate_file_hash(filepath: str) -> str:
    """Calcular hash SHA256 de un archivo"""
    sha256 = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b''):
            sha256.update(chunk)
    return sha256.hexdigest()


def calculate_db_hash() -> str:
    """Calcular hash SHA256 de la base de datos"""
    return calculate_file_hash(DB_PATH)


def create_backup(backup_type: str = 'manual', description: str = None) -> Dict:
    """Crear backup de la base de datos"""
    import shutil
    
    os.makedirs(BACKUP_DIR, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"chingin_backup_{timestamp}.db"
    backup_filepath = os.path.join(BACKUP_DIR, backup_filename)
    
    # Copiar base de datos
    shutil.copy2(DB_PATH, backup_filepath)
    
    # Calcular hash
    file_hash = calculate_file_hash(backup_filepath)
    file_size = os.path.getsize(backup_filepath)
    
    # Registrar backup
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO backups (filename, filepath, file_hash, file_size, backup_type, description)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (backup_filename, backup_filepath, file_hash, file_size, backup_type, description))
    
    # Log de auditoría
    log_audit('CREATE_BACKUP', 'backups', backup_filename, None, None, 
              f"Hash: {file_hash}, Size: {file_size}")
    
    # Limpiar backups antiguos
    cleanup_old_backups()
    
    return {
        'filename': backup_filename,
        'filepath': backup_filepath,
        'hash': file_hash,
        'size': file_size,
        'created_at': timestamp
    }


def verify_backup_integrity(backup_id: int) -> Dict:
    """Verificar integridad de un backup"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM backups WHERE id = ?", (backup_id,))
        backup = cursor.fetchone()
        
        if not backup:
            return {'valid': False, 'error': 'Backup no encontrado'}
        
        backup = dict(backup)
        
        if not os.path.exists(backup['filepath']):
            return {'valid': False, 'error': 'Archivo no existe'}
        
        current_hash = calculate_file_hash(backup['filepath'])
        is_valid = current_hash == backup['file_hash']
        
        # Actualizar estado
        cursor.execute("""
            UPDATE backups SET is_valid = ? WHERE id = ?
        """, (1 if is_valid else 0, backup_id))
        
        return {
            'valid': is_valid,
            'stored_hash': backup['file_hash'],
            'current_hash': current_hash,
            'filename': backup['filename']
        }


def restore_from_backup(backup_id: int) -> Dict:
    """Restaurar base de datos desde backup"""
    import shutil
    
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM backups WHERE id = ?", (backup_id,))
        backup = cursor.fetchone()
        
        if not backup:
            return {'success': False, 'error': 'Backup no encontrado'}
        
        backup = dict(backup)
    
    # Verificar integridad primero
    integrity = verify_backup_integrity(backup_id)
    if not integrity['valid']:
        return {'success': False, 'error': f"Backup corrupto: {integrity.get('error', 'Hash no coincide')}"}
    
    # Crear backup del estado actual antes de restaurar
    current_backup = create_backup('pre_restore', 'Backup antes de restauración')
    
    # Restaurar
    try:
        shutil.copy2(backup['filepath'], DB_PATH)
        log_audit('RESTORE_BACKUP', 'backups', str(backup_id), None, None,
                  f"Restaurado desde: {backup['filename']}")
        
        return {
            'success': True,
            'restored_from': backup['filename'],
            'previous_backup': current_backup['filename']
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


def get_backups() -> List[Dict]:
    """Obtener lista de backups"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM backups ORDER BY created_at DESC")
        return [dict(row) for row in cursor.fetchall()]


def cleanup_old_backups():
    """Eliminar backups antiguos"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Obtener configuración
        cursor.execute("SELECT value FROM settings WHERE key = 'max_backups_keep'")
        row = cursor.fetchone()
        max_keep = int(row[0]) if row else 30
        
        # Obtener backups a eliminar
        cursor.execute("""
            SELECT id, filepath FROM backups 
            ORDER BY created_at DESC
            LIMIT -1 OFFSET ?
        """, (max_keep,))
        
        old_backups = cursor.fetchall()
        
        for backup in old_backups:
            backup_id, filepath = backup
            if os.path.exists(filepath):
                os.remove(filepath)
            cursor.execute("DELETE FROM backups WHERE id = ?", (backup_id,))
        
        if old_backups:
            log_audit('CLEANUP_BACKUPS', 'backups', None, None, None,
                      f"Eliminados {len(old_backups)} backups antiguos")


def check_auto_backup():
    """Verificar si se necesita backup automático"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        # Verificar si está habilitado
        cursor.execute("SELECT value FROM settings WHERE key = 'auto_backup_enabled'")
        row = cursor.fetchone()
        if not row or row[0] != 'true':
            return None
        
        # Obtener intervalo
        cursor.execute("SELECT value FROM settings WHERE key = 'auto_backup_interval_hours'")
        row = cursor.fetchone()
        interval_hours = int(row[0]) if row else 24
        
        # Obtener último backup automático
        cursor.execute("""
            SELECT created_at FROM backups 
            WHERE backup_type = 'auto' 
            ORDER BY created_at DESC LIMIT 1
        """)
        row = cursor.fetchone()
        
        if not row:
            # No hay backups, crear uno
            return create_backup('auto', 'Backup automático inicial')
        
        last_backup = datetime.fromisoformat(row[0].replace('Z', '+00:00') if 'Z' in row[0] else row[0])
        if datetime.now() - last_backup > timedelta(hours=interval_hours):
            return create_backup('auto', f'Backup automático ({interval_hours}h)')
        
        return None


# ========================================
# FUNCIONES DE CONFIGURACIÓN
# ========================================

def get_setting(key: str) -> Optional[str]:
    """Obtener configuración"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
        row = cursor.fetchone()
        return row[0] if row else None


def set_setting(key: str, value: str, description: str = None):
    """Guardar configuración"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO settings (key, value, description, updated_at)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(key) DO UPDATE SET
                value = excluded.value,
                updated_at = CURRENT_TIMESTAMP
        """, (key, value, description))


def get_all_settings() -> Dict:
    """Obtener todas las configuraciones"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT key, value, description FROM settings")
        return {row[0]: {'value': row[1], 'description': row[2]} for row in cursor.fetchall()}


# ========================================
# FUNCIONES DE SINCRONIZACIÓN DE EMPLEADOS
# ========================================

# Nota: El nombre del archivo tiene un espacio japonés (全角スペース \u3000) entre T y 2022
# Usar formato Unix-like para compatibilidad con bash en Windows
import os as _os
MASTER_DIR = "//UNS-Kikaku/共有フォルダ/SCTDateBase"
MASTER_FILENAME = "【新】社員台帳(UNS)T\u30002022.04.05～.xlsm"
MASTER_FILE_PATH = MASTER_DIR + "/" + MASTER_FILENAME


def sync_haken_employees() -> Dict:
    """Sincronizar empleados 派遣社員 desde Excel maestro"""
    import openpyxl
    
    print(f"DEBUG: MASTER_FILE_PATH = {repr(MASTER_FILE_PATH)}")
    print(f"DEBUG: File exists = {_os.path.exists(MASTER_FILE_PATH)}")
    
    try:
        wb = openpyxl.load_workbook(MASTER_FILE_PATH, read_only=True, data_only=True)
        ws = wb['DBGenzaiX']
        
        count_inserted = 0
        count_updated = 0
        
        with get_connection() as conn:
            cursor = conn.cursor()
            
            # Leer desde fila 2 (la fila 1 son headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]:  # Si no hay 社員№, saltar
                    continue
                
                employee_id = str(row[1]).strip()
                
                # Formatear fechas
                def format_date(val):
                    if val is None:
                        return None
                    if hasattr(val, 'strftime'):
                        return val.strftime('%Y-%m-%d')
                    return str(val)
                
                data = {
                    'employee_id': employee_id,
                    'status': str(row[0]) if row[0] else None,
                    'dispatch_id': str(row[2]) if row[2] else None,
                    'dispatch_company': str(row[3]) if row[3] else None,
                    'department': str(row[4]) if row[4] else None,
                    'line': str(row[5]) if row[5] else None,
                    'job_description': str(row[6]) if row[6] else None,
                    'name': str(row[7]) if row[7] else None,
                    'name_kana': str(row[8]) if row[8] else None,
                    'gender': str(row[9]) if row[9] else None,
                    'nationality': str(row[10]) if row[10] else None,
                    'birth_date': format_date(row[11]),
                    'age': int(row[12]) if row[12] else None,
                    'hourly_rate': float(row[13]) if row[13] else None,
                    'hourly_rate_history': str(row[14]) if row[14] else None,
                    'billing_rate': float(row[15]) if row[15] else None,
                    'billing_history': str(row[16]) if row[16] else None,
                    'profit_margin': float(row[17]) if row[17] else None,
                    'standard_salary': float(row[18]) if row[18] else None,
                    'health_insurance': float(row[19]) if row[19] else None,
                    'care_insurance': float(row[20]) if row[20] and row[20] != '0' else 0,
                    'pension': float(row[21]) if row[21] else None,
                    'visa_expiry': format_date(row[22]),
                    'visa_alert': str(row[23]) if row[23] else None,
                    'visa_type': str(row[24]) if row[24] else None,
                    'postal_code': str(row[25]) if row[25] else None,
                    'address': str(row[26]) if row[26] else None,
                    'apartment': str(row[27]) if row[27] else None,
                    'move_in_date': str(row[28]) if row[28] else None,
                    'hire_date': format_date(row[29]),
                }
                
                # Check if exists
                cursor.execute("SELECT id FROM haken_employees WHERE employee_id = ?", (employee_id,))
                exists = cursor.fetchone()
                
                if exists:
                    cursor.execute("""
                        UPDATE haken_employees SET
                            status=?, dispatch_id=?, dispatch_company=?, department=?,
                            line=?, job_description=?, name=?, name_kana=?, gender=?,
                            nationality=?, birth_date=?, age=?, hourly_rate=?,
                            hourly_rate_history=?, billing_rate=?, billing_history=?,
                            profit_margin=?, standard_salary=?, health_insurance=?,
                            care_insurance=?, pension=?, visa_expiry=?, visa_alert=?,
                            visa_type=?, postal_code=?, address=?, apartment=?,
                            move_in_date=?, hire_date=?, synced_at=CURRENT_TIMESTAMP
                        WHERE employee_id=?
                    """, (
                        data['status'], data['dispatch_id'], data['dispatch_company'],
                        data['department'], data['line'], data['job_description'],
                        data['name'], data['name_kana'], data['gender'], data['nationality'],
                        data['birth_date'], data['age'], data['hourly_rate'],
                        data['hourly_rate_history'], data['billing_rate'], data['billing_history'],
                        data['profit_margin'], data['standard_salary'], data['health_insurance'],
                        data['care_insurance'], data['pension'], data['visa_expiry'],
                        data['visa_alert'], data['visa_type'], data['postal_code'],
                        data['address'], data['apartment'], data['move_in_date'],
                        data['hire_date'], employee_id
                    ))
                    count_updated += 1
                else:
                    cursor.execute("""
                        INSERT INTO haken_employees (
                            employee_id, status, dispatch_id, dispatch_company, department,
                            line, job_description, name, name_kana, gender, nationality,
                            birth_date, age, hourly_rate, hourly_rate_history, billing_rate,
                            billing_history, profit_margin, standard_salary, health_insurance,
                            care_insurance, pension, visa_expiry, visa_alert, visa_type,
                            postal_code, address, apartment, move_in_date, hire_date
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        data['employee_id'], data['status'], data['dispatch_id'],
                        data['dispatch_company'], data['department'], data['line'],
                        data['job_description'], data['name'], data['name_kana'],
                        data['gender'], data['nationality'], data['birth_date'],
                        data['age'], data['hourly_rate'], data['hourly_rate_history'],
                        data['billing_rate'], data['billing_history'], data['profit_margin'],
                        data['standard_salary'], data['health_insurance'], data['care_insurance'],
                        data['pension'], data['visa_expiry'], data['visa_alert'],
                        data['visa_type'], data['postal_code'], data['address'],
                        data['apartment'], data['move_in_date'], data['hire_date']
                    ))
                    count_inserted += 1
        
        wb.close()
        
        # Log audit
        log_audit('SYNC_HAKEN', 'haken_employees', None, None, None,
                   json.dumps({'inserted': count_inserted, 'updated': count_updated}))
        
        return {
            'success': True,
            'inserted': count_inserted,
            'updated': count_updated,
            'total': count_inserted + count_updated
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


def sync_ukeoi_employees() -> Dict:
    """Sincronizar empleados 請負社員 desde Excel maestro"""
    import openpyxl
    
    try:
        wb = openpyxl.load_workbook(MASTER_FILE_PATH, read_only=True, data_only=True)
        ws = wb['DBUkeoiX']
        
        count_inserted = 0
        count_updated = 0
        
        with get_connection() as conn:
            cursor = conn.cursor()
            
            # Leer desde fila 2 (la fila 1 son headers)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[1]:  # Si no hay 社員№, saltar
                    continue
                
                employee_id = str(row[1]).strip()
                
                # Formatear fechas
                def format_date(val):
                    if val is None:
                        return None
                    if hasattr(val, 'strftime'):
                        return val.strftime('%Y-%m-%d')
                    # Podría ser un número serial de Excel
                    if isinstance(val, (int, float)):
                        try:
                            from datetime import datetime, timedelta
                            base_date = datetime(1899, 12, 30)
                            return (base_date + timedelta(days=int(val))).strftime('%Y-%m-%d')
                        except:
                            return str(val)
                    return str(val)
                
                data = {
                    'employee_id': employee_id,
                    'status': str(row[0]) if row[0] else None,
                    'job_type': str(row[2]) if row[2] else None,
                    'name': str(row[3]) if row[3] else None,
                    'name_kana': str(row[4]) if row[4] else None,
                    'gender': str(row[5]) if row[5] else None,
                    'nationality': str(row[6]) if row[6] else None,
                    'birth_date': format_date(row[7]),
                    'age': int(row[8]) if row[8] else None,
                    'hourly_rate': float(row[9]) if row[9] else None,
                    'hourly_rate_history': str(row[10]) if row[10] else None,
                    'standard_salary': float(row[11]) if row[11] else None,
                    'health_insurance': float(row[12]) if row[12] else None,
                    'care_insurance': float(row[13]) if row[13] and row[13] != '0' else 0,
                    'pension': float(row[14]) if row[14] and row[14] != '0' else 0,
                    'commute_distance': float(row[15]) if row[15] else None,
                    'transport_fee': float(row[16]) if row[16] else None,
                    'profit_margin': float(row[17]) if row[17] else None,
                    'visa_expiry': format_date(row[18]) if row[18] and str(row[18]).strip() else None,
                    'visa_alert': str(row[19]) if row[19] else None,
                    'visa_type': str(row[20]) if row[20] else None,
                    'postal_code': str(row[21]) if row[21] else None,
                    'address': str(row[22]) if row[22] else None,
                    'apartment': str(row[23]) if row[23] else None,
                    'move_in_date': str(row[24]) if row[24] else None,
                    'hire_date': format_date(row[25]),
                    'resignation_date': format_date(row[26]) if row[26] else None,
                    'move_out_date': str(row[27]) if row[27] else None,
                    'social_insurance': str(row[28]) if row[28] else None,
                    'account_name': str(row[29]) if row[29] else None,
                }
                
                # Check if exists
                cursor.execute("SELECT id FROM ukeoi_employees WHERE employee_id = ?", (employee_id,))
                exists = cursor.fetchone()
                
                if exists:
                    cursor.execute("""
                        UPDATE ukeoi_employees SET
                            status=?, job_type=?, name=?, name_kana=?, gender=?,
                            nationality=?, birth_date=?, age=?, hourly_rate=?,
                            hourly_rate_history=?, standard_salary=?, health_insurance=?,
                            care_insurance=?, pension=?, commute_distance=?, transport_fee=?,
                            profit_margin=?, visa_expiry=?, visa_alert=?, visa_type=?,
                            postal_code=?, address=?, apartment=?, move_in_date=?,
                            hire_date=?, resignation_date=?, move_out_date=?,
                            social_insurance=?, account_name=?, synced_at=CURRENT_TIMESTAMP
                        WHERE employee_id=?
                    """, (
                        data['status'], data['job_type'], data['name'], data['name_kana'],
                        data['gender'], data['nationality'], data['birth_date'], data['age'],
                        data['hourly_rate'], data['hourly_rate_history'], data['standard_salary'],
                        data['health_insurance'], data['care_insurance'], data['pension'],
                        data['commute_distance'], data['transport_fee'], data['profit_margin'],
                        data['visa_expiry'], data['visa_alert'], data['visa_type'],
                        data['postal_code'], data['address'], data['apartment'],
                        data['move_in_date'], data['hire_date'], data['resignation_date'],
                        data['move_out_date'], data['social_insurance'], data['account_name'],
                        employee_id
                    ))
                    count_updated += 1
                else:
                    cursor.execute("""
                        INSERT INTO ukeoi_employees (
                            employee_id, status, job_type, name, name_kana, gender,
                            nationality, birth_date, age, hourly_rate, hourly_rate_history,
                            standard_salary, health_insurance, care_insurance, pension,
                            commute_distance, transport_fee, profit_margin, visa_expiry,
                            visa_alert, visa_type, postal_code, address, apartment,
                            move_in_date, hire_date, resignation_date, move_out_date,
                            social_insurance, account_name
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """, (
                        data['employee_id'], data['status'], data['job_type'], data['name'],
                        data['name_kana'], data['gender'], data['nationality'], data['birth_date'],
                        data['age'], data['hourly_rate'], data['hourly_rate_history'],
                        data['standard_salary'], data['health_insurance'], data['care_insurance'],
                        data['pension'], data['commute_distance'], data['transport_fee'],
                        data['profit_margin'], data['visa_expiry'], data['visa_alert'],
                        data['visa_type'], data['postal_code'], data['address'],
                        data['apartment'], data['move_in_date'], data['hire_date'],
                        data['resignation_date'], data['move_out_date'], data['social_insurance'],
                        data['account_name']
                    ))
                    count_inserted += 1
        
        wb.close()
        
        # Log audit
        log_audit('SYNC_UKEOI', 'ukeoi_employees', None, None, None,
                   json.dumps({'inserted': count_inserted, 'updated': count_updated}))
        
        return {
            'success': True,
            'inserted': count_inserted,
            'updated': count_updated,
            'total': count_inserted + count_updated
        }
    except Exception as e:
        return {'success': False, 'error': str(e)}


def sync_all_employees() -> Dict:
    """Sincronizar todos los empleados (派遣 y 請負)"""
    result_haken = sync_haken_employees()
    result_ukeoi = sync_ukeoi_employees()
    
    return {
        'success': result_haken.get('success', False) and result_ukeoi.get('success', False),
        'haken': result_haken,
        'ukeoi': result_ukeoi
    }


def get_haken_employee(employee_id: str) -> Optional[Dict]:
    """Obtener empleado 派遣 por ID"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM haken_employees WHERE employee_id = ?", (employee_id,))
        row = cursor.fetchone()
        return dict(row) if row else None


def get_ukeoi_employee(employee_id: str) -> Optional[Dict]:
    """Obtener empleado 請負 por ID"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM ukeoi_employees WHERE employee_id = ?", (employee_id,))
        row = cursor.fetchone()
        return dict(row) if row else None


def get_employee_master(employee_id: str) -> Optional[Dict]:
    """Buscar empleado en ambas tablas maestro"""
    # Primero buscar en 派遣
    emp = get_haken_employee(employee_id)
    if emp:
        emp['type'] = '派遣社員'
        return emp
    
    # Luego buscar en 請負
    emp = get_ukeoi_employee(employee_id)
    if emp:
        emp['type'] = '請負社員'
        return emp
    
    return None


def get_all_haken_employees() -> List[Dict]:
    """Obtener todos los empleados 派遣"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM haken_employees ORDER BY employee_id")
        return {'employees': [dict(row) for row in cursor.fetchall()]}


def get_all_ukeoi_employees() -> Dict:
    """
    Obtener todos los empleados 請負
    Primero busca en ukeoi_employees (master),
    si está vacía, busca en employees con IDs 03xxxx (de nómina)
    """
    with get_connection() as conn:
        cursor = conn.cursor()

        # Primero intentar obtener de la tabla master
        cursor.execute("SELECT * FROM ukeoi_employees ORDER BY employee_id")
        master_employees = [dict(row) for row in cursor.fetchall()]

        # Si hay empleados en master, usar esos
        if master_employees:
            return {'employees': master_employees}

        # Si no hay en master, buscar en employees con IDs 03xxxx (請負社員 de nómina)
        cursor.execute("""
            SELECT DISTINCT
                employee_id,
                name_jp as name,
                name_jp as name_kana,
                name_roman,
                NULL as dispatch_company,
                NULL as job_type,
                NULL as hire_date,
                NULL as gender,
                'active' as status
            FROM employees
            WHERE employee_id LIKE '03%'
            ORDER BY employee_id
        """)
        payroll_employees = [dict(row) for row in cursor.fetchall()]

        return {'employees': payroll_employees}


def get_employee_master_stats() -> Dict:
    """Obtener estadísticas de empleados maestro"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        stats = {}
        
        # 派遣社員
        cursor.execute("SELECT COUNT(*) FROM haken_employees")
        stats['haken_total'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM haken_employees WHERE status = '在職中' OR status = '現在'")
        stats['haken_active'] = cursor.fetchone()[0]
        cursor.execute("SELECT MAX(synced_at) FROM haken_employees")
        row = cursor.fetchone()
        stats['haken_last_sync'] = row[0] if row else None
        
        # 請負社員
        cursor.execute("SELECT COUNT(*) FROM ukeoi_employees")
        stats['ukeoi_total'] = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM ukeoi_employees WHERE status = '在職中' OR status = '現在'")
        stats['ukeoi_active'] = cursor.fetchone()[0]
        cursor.execute("SELECT MAX(synced_at) FROM ukeoi_employees")
        row = cursor.fetchone()
        stats['ukeoi_last_sync'] = row[0] if row else None
        
        return stats


def get_dispatch_companies() -> Dict:
    """Obtener lista de派遣先 (fábricas/empresas) únicas"""
    with get_connection() as conn:
        cursor = conn.cursor()
        # Obtener派遣先 de empleados activos con conteo
        cursor.execute("""
            SELECT dispatch_company, COUNT(*) as employee_count
            FROM haken_employees 
            WHERE dispatch_company IS NOT NULL 
              AND dispatch_company != ''
              AND (status = '在職中' OR status = '現在')
            GROUP BY dispatch_company
            ORDER BY dispatch_company
        """)
        companies = [{'name': row[0], 'count': row[1]} for row in cursor.fetchall()]
        return {'companies': companies, 'total': len(companies)}


def get_ukeoi_job_types() -> Dict:
    """Obtener lista de 請負業務 (tipos de trabajo) únicos"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT job_type, COUNT(*) as employee_count
            FROM ukeoi_employees 
            WHERE job_type IS NOT NULL 
              AND job_type != ''
              AND (status = '在職中' OR status = '現在')
            GROUP BY job_type
            ORDER BY job_type
        """)
        job_types = [{'name': row[0], 'count': row[1]} for row in cursor.fetchall()]
        return {'job_types': job_types, 'total': len(job_types)}


def get_employees_by_company(company_name: str) -> Dict:
    """Obtener empleados派遣 de una fábrica específica"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT employee_id, name, name_kana, gender, status, hire_date
            FROM haken_employees 
            WHERE dispatch_company = ?
              AND (status = '在職中' OR status = '現在')
            ORDER BY name
        """, (company_name,))
        employees = [{'id': row[0], 'name': row[1], 'name_kana': row[2], 'gender': row[3]} for row in cursor.fetchall()]
        return {'employees': employees, 'count': len(employees)}


def get_employees_by_job_type(job_type: str) -> Dict:
    """Obtener empleados 請負 de un tipo de trabajo específico"""
    with get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT employee_id, name, name_kana, gender, status, hire_date
            FROM ukeoi_employees 
            WHERE job_type = ?
              AND (status = '在職中' OR status = '現在')
            ORDER BY name
        """, (job_type,))
        employees = [{'id': row[0], 'name': row[1], 'name_kana': row[2], 'gender': row[3]} for row in cursor.fetchall()]
        return {'employees': employees, 'count': len(employees)}


# ========================================
# ESTADÍSTICAS
# ========================================

def get_statistics() -> Dict:
    """Obtener estadísticas generales"""
    with get_connection() as conn:
        cursor = conn.cursor()
        
        stats = {}
        
        # Total empleados
        cursor.execute("SELECT COUNT(*) FROM employees WHERE status = 'active'")
        stats['total_employees'] = cursor.fetchone()[0]
        
        # Total registros de nómina
        cursor.execute("SELECT COUNT(*) FROM payroll_records")
        stats['total_payroll_records'] = cursor.fetchone()[0]
        
        # Periodos únicos
        cursor.execute("SELECT COUNT(DISTINCT period) FROM payroll_records")
        stats['total_periods'] = cursor.fetchone()[0]
        
        # Totales
        cursor.execute("SELECT SUM(total_pay), SUM(net_pay) FROM payroll_records")
        row = cursor.fetchone()
        stats['total_gross_pay'] = row[0] or 0
        stats['total_net_pay'] = row[1] or 0
        
        # Archivos procesados
        cursor.execute("SELECT COUNT(*) FROM processed_files WHERE status = 'success'")
        stats['files_processed'] = cursor.fetchone()[0]
        
        # Backups
        cursor.execute("SELECT COUNT(*) FROM backups WHERE is_valid = 1")
        stats['valid_backups'] = cursor.fetchone()[0]
        
        # Último backup
        cursor.execute("SELECT created_at FROM backups ORDER BY created_at DESC LIMIT 1")
        row = cursor.fetchone()
        stats['last_backup'] = row[0] if row else None
        
        # Integridad de BD
        stats['db_hash'] = calculate_db_hash()
        
        return stats


# ========================================
# INICIALIZACIÓN
# ========================================

if __name__ == "__main__":
    print("="*60)
    print("🗄️  Inicializando Base de Datos - ChinginApp v4 PRO")
    print("="*60)
    
    init_database()
    
    # Mostrar estadísticas
    stats = get_statistics()
    print(f"\n📊 Estadísticas:")
    print(f"   Empleados: {stats['total_employees']}")
    print(f"   Registros de nómina: {stats['total_payroll_records']}")
    print(f"   Periodos: {stats['total_periods']}")
    print(f"   Backups válidos: {stats['valid_backups']}")
    print(f"   Hash BD: {stats['db_hash'][:16]}...")
    
    print("\n✅ Base de datos lista!")
