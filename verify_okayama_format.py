#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import load_workbook

file_path = r"C:\Users\JPUNS\Desktop\Crear Chingi 25.11.17 docs\JPNuevo\給与明細(派遣社員)2025.1(0217支給).xlsm"

wb = load_workbook(file_path, data_only=True)

print("=" * 100)
print("VERIFICACIÓN: Hoja 岡山工場")
print("=" * 100)

ws = wb['岡山工場']
print(f"\nDimensiones: {ws.max_row} filas x {ws.max_column} columnas")

print(f"\nHeaders (fila 1, primeras 10 columnas):")
for col in range(1, 11):
    h = ws.cell(row=1, column=col).value
    print(f"  Col {col:2d}: {h}")

print(f"\nPrimer registro (fila 2, primeras 10 columnas):")
for col in range(1, 11):
    v = ws.cell(row=2, column=col).value
    print(f"  Col {col:2d}: {v}")

# Buscar empleados 03xxxx
print(f"\nBuscando empleados con ID 03xxxx:")
count_03 = 0
count_other = 0

for row in range(2, min(ws.max_row + 1, 100)):
    emp_id = ws.cell(row=row, column=2).value  # Col 2 = 従業員番号
    if emp_id:
        emp_id_str = str(emp_id).strip()
        name = ws.cell(row=row, column=4).value

        if emp_id_str.startswith('03'):
            count_03 += 1
            if count_03 <= 5:
                print(f"  ✅ 請負 - Fila {row}: ID={emp_id}, Nombre={name}")
        else:
            count_other += 1
            if count_other <= 2:
                print(f"  ⚪ Otro - Fila {row}: ID={emp_id}, Nombre={name}")

print(f"\nResumen (primeras 98 filas):")
print(f"  Empleados 03xxxx (請負社員): {count_03}")
print(f"  Otros empleados: {count_other}")

print("\n" + "=" * 100)
print("CONCLUSIÓN:")
print("=" * 100)
if count_03 > 0:
    print(f"✅ La hoja 岡山工場 TIENE empleados 請負社員 (03xxxx)")
    print(f"✅ Formato es tabular igual que totalChin")
    print(f"\n💡 SOLUCIÓN: Agregar '岡山工場' a la lista de hojas prioritarias")
else:
    print(f"❌ La hoja 岡山工場 NO tiene empleados 請負社員")

wb.close()
